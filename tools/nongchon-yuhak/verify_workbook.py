"""만든 워크북의 수식을 전부 계산해서 값이 맞는지 확인한다.

  python verify_workbook.py <워크북.xlsx> <원본.xlsx>

LibreOffice 로 재계산할 수 없는 환경이 있어서 formula_eval.py 로 직접 계산한다.
계산이 되는지만 보지 않고, 집계 시트가 낸 숫자를 원데이터에서 따로 센 값과
맞춰 본다.
"""

from __future__ import annotations

import re
import sys

import openpyxl
from openpyxl.utils import get_column_letter

import formula_eval as F
import transform as T

# 엑셀 2007 이전 함수만 쓴다. 이후 함수는 접두사 문제로 다른 프로그램에서 깨진다.
ALLOWED = {
    "COUNTIF", "COUNTIFS", "COUNTA", "COUNTBLANK", "SUM", "IF", "OR", "AND",
    "MIN", "MAX", "INDEX", "MATCH", "IFERROR", "SUMPRODUCT", "SUMIF", "SUMIFS",
}
BANNED = {"XLOOKUP", "XMATCH", "SORT", "FILTER", "UNIQUE", "SEQUENCE",
          "TEXTJOIN", "CONCAT", "IFS", "SWITCH", "MAXIFS", "MINIFS"}

FUNC_RE = re.compile(r"([A-Z_][A-Z0-9_.]*)\s*\(")


def all_formulas(wb):
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    yield ws.title, c.coordinate, c.value


def check_syntax(wb):
    problems = []
    for sheet, coord, text in all_formulas(wb):
        fns = {m.group(1) for m in FUNC_RE.finditer(text.upper())}
        if fns & BANNED:
            problems.append(f"{sheet}!{coord}: 다른 프로그램에서 깨지는 함수 {sorted(fns & BANNED)}")
        if fns - ALLOWED - BANNED:
            problems.append(f"{sheet}!{coord}: 확인이 필요한 함수 {sorted(fns - ALLOWED - BANNED)}")
        if text.count("(") != text.count(")"):
            problems.append(f"{sheet}!{coord}: 괄호가 맞지 않음 — {text[:70]}")
        for m in re.finditer(r"'([^']+)'!", text):
            if m.group(1) not in wb.sheetnames:
                problems.append(f"{sheet}!{coord}: 없는 시트 '{m.group(1)}'")
        # 머리글을 '=' 로 시작하게 적으면 엑셀이 수식으로 읽는다
        if not FUNC_RE.search(text.upper()) and not re.search(r"[A-Z]\$?\d", text):
            problems.append(f"{sheet}!{coord}: 수식이 아닌 글인데 '=' 로 시작함 — {text[:40]}")
    return problems


def next_term(e, present):
    """유학이력 '다음 학기' 칸이 내야 하는 값 — 워크북 수식과 같은 순서로 본다."""
    if e.status == "예정":
        return "예정"
    if not (e.end_date is None or e.end_date >= e.term_end):
        return "학기 중 종료"
    if (e.student_id, T.sem_index(e.year, e.term) + 1) in present:
        return "이어짐"
    return "학기말 종료" if e.status == "종료" else "미정"


def expected(data):
    """원데이터에서 따로 센 값."""
    total, stay, region, school = {}, {}, {}, {}
    flow = {}
    present = {(e.student_id, T.sem_index(e.year, e.term)) for e in data["enrollments"]}
    for e in data["enrollments"]:
        lab = T.sem_label(e.year, e.term)
        total[lab] = total.get(lab, 0) + 1
        if e.end_date is None or e.end_date >= e.term_end:
            stay[lab] = stay.get(lab, 0) + 1
        region[(lab, e.region)] = region.get((lab, e.region), 0) + 1
        school[(lab, e.school)] = school.get((lab, e.school), 0) + 1
        kind = "신규" if e.kind == "신규" else "계속·재유학"
        flow[(lab, kind)] = flow.get((lab, kind), 0) + 1
        nx = next_term(e, present)
        if nx == "예정":
            nx = "미정"
        flow[(lab, nx)] = flow.get((lab, nx), 0) + 1
    return total, stay, region, school, flow


def header_col(ws, header, row=1):
    for c in range(1, ws.max_column + 1):
        if ws.cell(row, c).value == header:
            return get_column_letter(c)
    raise KeyError(f"{ws.title} 에 '{header}' 열이 없음")


def check_values(wb, book, data):
    problems, checked = [], 0
    total, stay, region, school, flow = expected(data)

    # 유학이력 '다음 학기' — 한 줄씩 전부 대조한다
    present = {(e.student_id, T.sem_index(e.year, e.term)) for e in data["enrollments"]}
    ws = wb["유학이력"]
    nxt_col = header_col(ws, "다음 학기")
    for i, e in enumerate(data["enrollments"]):
        r = i + 2
        checked += 1
        got = book.value("유학이력", f"{nxt_col}{r}")
        want = next_term(e, present)
        if got != want:
            problems.append(
                f"유학이력 {r}행 ({e.name} {T.sem_label(e.year, e.term)}) 다음 학기: "
                f"수식 {got} ≠ 원데이터 {want}"
            )

    # 학기별집계 — 들어온 쪽·인원·나간 쪽이 모두 맞고, 양쪽 합이 인원과 같아야 한다
    ws = wb["학기별집계"]
    FLOW_COLS = ((2, "신규"), (3, "계속·재유학"), (5, "이어짐"),
                 (6, "학기말 종료"), (7, "미정"), (8, "학기 중 종료"))
    for r in range(5, 5 + len(data["semesters"])):
        lab = ws.cell(r, 1).value
        for col, want, what in ((4, total, "유학생 수(공식 기준)"), (9, stay, "학기말 재적")):
            checked += 1
            got = book.value("학기별집계", f"{get_column_letter(col)}{r}")
            if int(got or 0) != want.get(lab, 0):
                problems.append(f"학기별집계 {lab} {what}: 수식 {got} ≠ 원데이터 {want.get(lab, 0)}")
        for col, what in FLOW_COLS:
            checked += 1
            got = book.value("학기별집계", f"{get_column_letter(col)}{r}")
            if int(got or 0) != flow.get((lab, what), 0):
                problems.append(
                    f"학기별집계 {lab} {what}: 수식 {got} ≠ 원데이터 {flow.get((lab, what), 0)}"
                )
        # 들어온 쪽 합·나간 쪽 합이 모두 인원과 같아야 한다
        for cols, what in (((2, 3), "신규+계속"), ((5, 6, 7, 8), "이어짐+종료+미정")):
            checked += 1
            got = sum(int(book.value("학기별집계", f"{get_column_letter(c)}{r}") or 0) for c in cols)
            if got != total.get(lab, 0):
                problems.append(
                    f"학기별집계 {lab} {what} 합계 {got} ≠ 유학생 수 {total.get(lab, 0)}"
                )

    # 학기별집계의 학년도 블록 — 한 칸씩 더하고 빼면 그대로 이어져야 하고,
    # 학년도 인원은 중복을 뺀 사람 수와 같아야 한다
    ws = wb["학기별집계"]
    hrow = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 12).value == "셈법")
    year_students, sem_students = {}, {}
    for e in data["enrollments"]:
        year_students.setdefault(e.year, set()).add(e.student_id)
        sem_students.setdefault((e.year, e.term), set()).add(e.student_id)
    r = hrow + 1
    while ws.cell(r, 1).value:
        y = int(str(ws.cell(r, 1).value)[:4])
        a = sem_students.get((y, 1), set())
        b = sem_students.get((y, 2), set())
        nxt = sem_students.get((y + 1, 1), set())
        for col, want, what in (
            (2, len(a), "1학기 인원"),
            (5, len(a & b), "2학기로 이어짐"),
            (6, len(b - a), "2학기 신규"),
            (7, len(b), "2학기 인원"),
            (10, len(b & nxt), "다음 학년도로"),
            (11, len(year_students.get(y, ())), "학년도 인원"),
        ):
            checked += 1
            got = book.value("학기별집계", f"{get_column_letter(col)}{r}")
            if int(got or 0) != want:
                problems.append(f"학기별집계 {y}학년도 {what}: 수식 {got} ≠ 원데이터 {want}")
        # 종료 + 미정 + 이어짐 이 그 학기 인원과 맞아야 흐름이 끊기지 않는다
        for parts, whole, what in (((3, 4, 5), 2, "1학기"), ((8, 9, 10), 7, "2학기")):
            checked += 1
            got = sum(int(book.value("학기별집계", f"{get_column_letter(c)}{r}") or 0)
                      for c in parts)
            want = int(book.value("학기별집계", f"{get_column_letter(whole)}{r}") or 0)
            if got != want:
                problems.append(
                    f"학기별집계 {y}학년도 {what} 종료+미정+이어짐 {got} ≠ 인원 {want}"
                )
        r += 1

    # 지역별현황 / 학교별현황 격자
    for sheet, key_col, want in (("지역별현황", 1, region), ("학교별현황", 2, school)):
        ws = wb[sheet]
        sems = []
        c = key_col + 1
        while isinstance(ws.cell(2, c).value, str) and "학기" in ws.cell(2, c).value:
            sems.append((get_column_letter(c), ws.cell(2, c).value))
            c += 1
        r = 3
        while True:
            key = ws.cell(r, key_col).value
            if not key or key == "합계":
                break
            for letter, lab in sems:
                checked += 1
                got = book.value(sheet, f"{letter}{r}")
                if int(got or 0) != want.get((lab, key), 0):
                    problems.append(
                        f"{sheet} {key} {lab}: 수식 {got} ≠ 원데이터 {want.get((lab, key), 0)}"
                    )
            r += 1
        # 합계 행
        for letter, lab in sems:
            checked += 1
            got = book.value(sheet, f"{letter}{r}")
            if int(got or 0) != total.get(lab, 0):
                problems.append(f"{sheet} 합계 {lab}: 수식 {got} ≠ 원데이터 {total.get(lab, 0)}")

    # 연도별현황 — 학년도로 묶은 값이 중복 없이 센 값과 같아야 한다
    year_region, year_school, year_total = {}, {}, {}
    for e in data["enrollments"]:
        year_region.setdefault((e.year, e.region), set()).add(e.student_id)
        year_school.setdefault((e.year, e.school), set()).add(e.student_id)
        year_total.setdefault(e.year, set()).add(e.student_id)
    ws = wb["연도별현황"]
    years = []
    c = 2
    while isinstance(ws.cell(4, c).value, str) and "학년도" in ws.cell(4, c).value:
        years.append((get_column_letter(c), int(ws.cell(4, c).value[:4])))
        c += 1
    r = 5
    while ws.cell(r, 1).value and ws.cell(r, 1).value != "합계":
        key = ws.cell(r, 1).value
        for letter, y in years:
            checked += 1
            got = book.value("연도별현황", f"{letter}{r}")
            want = len(year_region.get((y, key), ()))
            if int(got or 0) != want:
                problems.append(f"연도별현황 {key} {y}학년도: 수식 {got} ≠ 원데이터 {want}")
        r += 1
    for letter, y in years:
        checked += 1
        got = book.value("연도별현황", f"{letter}{r}")
        if int(got or 0) != len(year_total.get(y, ())):
            problems.append(f"연도별현황 합계 {y}학년도: 수식 {got} ≠ 원데이터 {len(year_total.get(y, ()))}")

    # 학교 블록
    head = r + 3
    school_years = []
    c = 3
    while isinstance(ws.cell(head, c).value, str) and "학년도" in ws.cell(head, c).value:
        school_years.append((get_column_letter(c), int(ws.cell(head, c).value[:4])))
        c += 1
    r = head + 1
    while ws.cell(r, 2).value and ws.cell(r, 1).value != "합계":
        key = ws.cell(r, 2).value
        for letter, y in school_years:
            checked += 1
            got = book.value("연도별현황", f"{letter}{r}")
            want = len(year_school.get((y, key), ()))
            if int(got or 0) != want:
                problems.append(f"연도별현황 학교 {key} {y}학년도: 수식 {got} ≠ 원데이터 {want}")
        r += 1

    # 서울원적 — 학기별 요약이 원데이터의 서울 재적과 같아야 한다
    seoul = {}
    for e in data["enrollments"]:
        if e.home_region == "서울":
            lab = T.sem_label(e.year, e.term)
            seoul[lab] = seoul.get(lab, 0) + 1
    ws = wb["서울원적"]
    c = 2
    while isinstance(ws.cell(4, c).value, str) and "학기" in ws.cell(4, c).value:
        lab = ws.cell(4, c).value
        checked += 1
        got = book.value("서울원적", f"{get_column_letter(c)}5")
        if int(got or 0) != seoul.get(lab, 0):
            problems.append(f"서울원적 {lab}: 수식 {got} ≠ 원데이터 {seoul.get(lab, 0)}")
        c += 1

    # 학생현황의 유학 학기수 합이 유학이력 행 수와 같아야 한다
    ws = wb["학생현황"]
    col = header_col(ws, "유학 학기수")
    s = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value is None:
            break
        s += int(book.value("학생현황", f"{col}{r}") or 0)
    checked += 1
    if s != len(data["enrollments"]):
        problems.append(f"학생현황 유학 학기수 합계 {s} ≠ 유학이력 {len(data['enrollments'])}행")

    # 학생현황의 '현재 단계' 는 학생 수와 맞아야 하고, 모집단계현황의 집계와도 같아야 한다
    ws = wb["학생현황"]
    stage_col = header_col(ws, "현재 단계")
    seen = {}
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value is None:
            break
        v = book.value("학생현황", f"{stage_col}{r}")
        seen[v] = seen.get(v, 0) + 1
    checked += 1
    if sum(seen.values()) != len(data["students"]):
        problems.append(
            f"학생현황 현재 단계 합계 {sum(seen.values())} ≠ 학생 {len(data['students'])}명"
        )
    if set(seen) - set(T.STAGES):
        problems.append(f"학생현황에 모르는 단계가 있음: {sorted(set(seen) - set(T.STAGES))}")

    ws = wb["모집단계현황"]
    for r in range(1, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if name in T.STAGES and ws.cell(r, 3).value:
            checked += 1
            got = book.value("모집단계현황", f"B{r}")
            if int(got or 0) != seen.get(name, 0):
                problems.append(
                    f"모집단계현황 '{name}': 수식 {got} ≠ 학생현황 {seen.get(name, 0)}"
                )

    return problems, checked


def main(book_path, source_path):
    wb = openpyxl.load_workbook(book_path)
    formulas = list(all_formulas(wb))
    syntax = check_syntax(wb)

    book = F.Book(wb)
    for sheet, coord, _ in formulas:
        book.value(sheet, coord)

    data = T.build(source_path)
    values, checked = check_values(wb, book, data)

    print(f"수식 {len(formulas)}개")
    print(f"  문법 검사   : {'통과' if not syntax else f'{len(syntax)}건 문제'}")
    for p in syntax[:15]:
        print("   !", p)
    print(f"  계산 오류   : {len(book.errors)}건")
    for e in book.errors[:15]:
        print("   !", e)
    print(f"  값 대조     : {checked}건 중 {len(values)}건 불일치")
    for p in values[:15]:
        print("   !", p)

    ok = not syntax and not book.errors and not values
    print("모두 통과" if ok else "확인 필요")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main(sys.argv[1], sys.argv[2]))
