"""정규화한 데이터를 관리용 엑셀 워크북으로 쓴다.

  python build_workbook.py <원본.xlsx> <출력.xlsx>

집계 시트는 모두 수식(COUNTIFS/SUMIFS)으로 쓴다. 원데이터 시트에 행을 더하면
집계·현황·체제비 시트가 다시 계산된다.
"""

from __future__ import annotations

import datetime as dt
import sys
from collections import Counter

import openpyxl
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import transform as T

FONT = "맑은 고딕"
INK = "1F3864"        # 제목·헤더
ACCENT = "2E5C8A"
HEAD_FILL = PatternFill("solid", fgColor=INK)
SUB_FILL = PatternFill("solid", fgColor="D9E2F3")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")   # 사용자가 고치는 칸
DERIVED_FILL = PatternFill("solid", fgColor="F2F2F2")  # 수식 칸
BAND = PatternFill("solid", fgColor="F7F9FC")
HAIR = Side(style="thin", color="BFBFBF")
BOX = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)

DATE_FMT = "yyyy-mm-dd"

# 데이터 시트의 수식이 참조할 범위(여유분 포함). 행을 더 넣어도 집계가 따라온다.
LIMIT_STUDENT = 1200
LIMIT_APP = 1200
LIMIT_ENROLL = 2000


def col_of(headers: list[str], name: str) -> str:
    return get_column_letter(headers.index(name) + 1)


def rng(sheet: str, headers: list[str], name: str, limit: int) -> str:
    c = col_of(headers, name)
    return f"'{sheet}'!${c}$2:${c}${limit}"


def write_table(ws, headers, rows, widths, *, date_cols=(), wrap_cols=(), band=True,
                start_row=1, freeze="C"):
    """헤더 한 줄 + 데이터. 틀 고정과 필터까지 걸어둔다."""
    for i, h in enumerate(headers, start=1):
        c = ws.cell(start_row, i, h)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
    ws.row_dimensions[start_row].height = 34

    for r, row in enumerate(rows, start=start_row + 1):
        for i, v in enumerate(row, start=1):
            c = ws.cell(r, i, v)
            c.font = Font(name=FONT, size=10)
            c.border = BOX
            c.alignment = Alignment(
                vertical="center",
                wrap_text=headers[i - 1] in wrap_cols,
                horizontal="left" if headers[i - 1] in wrap_cols else None,
            )
            if headers[i - 1] in date_cols:
                c.number_format = DATE_FMT
                c.alignment = Alignment(horizontal="center", vertical="center")
            if band and (r - start_row) % 2 == 0:
                c.fill = BAND

    for name, w in widths.items():
        ws.column_dimensions[col_of(headers, name)].width = w
    ws.freeze_panes = f"{freeze}{start_row + 1}"
    last = max(start_row + len(rows), start_row + 1)
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{last}"
    return last


def add_dropdown(ws, headers, name, options, last_row):
    dv = DataValidation(
        type="list", formula1='"' + ",".join(options) + '"', allow_blank=True, showErrorMessage=False
    )
    ws.add_data_validation(dv)
    c = col_of(headers, name)
    dv.add(f"{c}2:{c}{last_row}")


# ─────────────────────────────────────────────────────────── 안내


def sheet_guide(ws, data, src_name):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 96

    r = [2]

    def title(text):
        c = ws.cell(r[0], 2, text)
        c.font = Font(name=FONT, size=16, bold=True, color=INK)
        ws.merge_cells(start_row=r[0], start_column=2, end_row=r[0], end_column=3)
        ws.row_dimensions[r[0]].height = 24
        r[0] += 2

    def head(text):
        c = ws.cell(r[0], 2, text)
        c.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=ACCENT)
        c.alignment = Alignment(vertical="center", indent=1)
        c2 = ws.cell(r[0], 3, "")
        c2.fill = PatternFill("solid", fgColor=ACCENT)
        ws.row_dimensions[r[0]].height = 20
        r[0] += 1

    def line(label, text):
        a = ws.cell(r[0], 2, label)
        a.font = Font(name=FONT, size=10, bold=True, color=INK)
        a.alignment = Alignment(vertical="top", indent=1)
        b = ws.cell(r[0], 3, text)
        b.font = Font(name=FONT, size=10)
        b.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
        ws.row_dimensions[r[0]].height = 15 + 13 * (len(text) // 60)
        r[0] += 1

    def gap():
        r[0] += 1

    e = data["enrollments"]
    apps = data["applications"]
    title("전북 농어촌유학생 관리 워크북")
    line("만든 날", dt.date.today().isoformat())
    line("원본 파일", src_name)
    line(
        "원본 범위",
        f"1번 시트 '{T.SOURCE_SHEET}' 만 읽었습니다. "
        "숨김 시트(연도별 명단, 정보공개 등)는 과거 산출물이라 가져오지 않았습니다.",
    )
    line(
        "기준 학기",
        f"{T.sem_label(T.CURRENT_YEAR, T.CURRENT_TERM)}. "
        "원본 종료일 칸의 '현재' 는 이 학기까지 재학 중이라는 뜻으로 풀었습니다.",
    )
    gap()

    head("무엇이 달라졌나")
    line(
        "한 줄 요약",
        "한 시트에 쌓여 있던 741행을 성격이 다른 세 개의 표로 나누고, 자주 뽑던 숫자는 수식으로 미리 걸어 두었습니다.",
    )
    line(
        "쪼갠 이유",
        "원본은 1행이 '한 번의 신청'이었는데, 그 안에 학생 정보·심사 결과·여러 해에 걸친 유학 이력이 함께 들어 있었습니다. "
        "그래서 '2025학년도 1학기 몇 명'을 세려면 유학 학년도 칸을 검색하고 손으로 빼는 수밖에 없었습니다.",
    )
    gap()

    head("시트 구성")
    line(
        "학생현황",
        f"1행 = 학생 1명 ({len(data['students'])}명). 총괄 시트입니다. "
        "접수·가배정·참가신청서·최종배정·유학 시작·유학 종료 칸을 옆으로 늘어놓고, "
        "'현재 단계' 는 그 칸들을 뒤에서부터 보고 수식이 정합니다. 칸을 채우면 단계가 따라 바뀝니다.",
    )
    line(
        "모집단계현황",
        "접수한 학생이 단계마다 얼마나 남는지, 그리고 지금 학생들이 어느 단계에 몇 명씩 있는지.",
    )
    line(
        "유학이력",
        f"1행 = 학생 × 학기 ({len(e)}행). 실제로 유학한 학기를 한 줄씩 펼친 표로, 모든 집계의 바탕입니다.",
    )
    line(
        "신청이력",
        f"1행 = 원본 1행 = 모집 차수별 신청 건 ({len(apps)}행). 원본 열을 모두 그대로 두고 앞쪽에 정리 열만 붙였습니다.",
    )
    line("학기별집계", "학년도·학기별 인원, 신규/계속, 성별, 거주 유형, 운영 지역·학교 수. 전부 수식입니다.")
    line("연도별현황", "시군별·학교별 인원을 학년도로 묶은 표. 발표 자료와 같은 기준입니다.")
    line(
        "서울원적_제출용",
        "서울시교육청이 준 서식 그대로 만든 제출용 명단. 왼쪽 15개 열이 서식이고, "
        "노란 열은 검토용이라 제출할 때 지웁니다. 지원금 대상 여부(○/×)는 최초 참여 학기부터 "
        "6학기 이내인지로 계산합니다.",
    )
    line(
        "서울원적",
        "원 지역이 서울인 학생만 모은 표. 서울시교육청 유학경비 지원 보고용입니다. "
        "학기 칸의 'O' 로 필요한 학기를 걸러 씁니다. 서울시교육청 양식을 받으면 열 차례를 맞춥니다.",
    )
    line("지역별현황 / 학교별현황", "같은 내용을 학기별로 늘어놓은 표.")
    line("체제비관리", "학생별 지원 시작 학기와 도청 3년 지원 만료 시점, 잔여 학기.")
    line("설정", "기준일과 체제비 단가. 지침이 바뀌면 노란 칸만 고치면 됩니다.")
    line(
        "데이터검증",
        f"확인 항목 {len(data['issues'])}건. 맨 위에 발표 자료와의 대조표가 있고, "
        f"그 아래 목록은 '조치 필요'({sum(1 for i in data['issues'] if i['심각도'] == '조치 필요')}건)와 "
        f"'참고'({sum(1 for i in data['issues'] if i['심각도'] == '참고')}건)로 나뉩니다. "
        "'참고'는 집계 숫자에 영향이 없습니다.",
    )
    line("원본_전체", "원본 1번 시트를 값 그대로 옮겨 둔 사본. 대조용이며 고치지 않습니다.")
    gap()

    head("고칠 때")
    line("노란 칸", "직접 입력하는 칸입니다.")
    line("회색 칸", "수식으로 채워지는 칸입니다. 값을 덮어쓰면 계산이 끊깁니다.")
    line(
        "새 유학생이 오면",
        "① 학생현황에 학생을 한 줄 추가하고(학생ID는 S0699처럼 이어서) "
        "② 신청이력에 신청 건을 한 줄 "
        "③ 유학이력에 학기당 한 줄씩 추가합니다. 집계 시트는 자동으로 따라옵니다.",
    )
    line("학기가 바뀌면", "계속 유학하는 학생은 유학이력에 새 학기 줄을 한 줄씩 더합니다.")
    line("수식 범위", f"학생현황 {LIMIT_STUDENT}행 / 신청이력 {LIMIT_APP}행 / 유학이력 {LIMIT_ENROLL}행까지 미리 잡아 두었습니다.")
    gap()

    head("숫자를 어떻게 셌나")
    line(
        "학기 나누는 법",
        "3월 1일 시작이 1학기, 9월 1일 시작이 2학기입니다. 학년도는 3월에 바뀝니다. "
        "예를 들어 2022년 10월 시작은 2022학년도 2학기입니다.",
    )
    line(
        "접수일 칸 읽는 법",
        "원본 '농어촌유학 접수일/시작일' 칸의 마지막 자리는 날짜가 아니라 모집 차수입니다. "
        "2025-09-02는 2025학년도 2학기 2차 모집이라는 뜻입니다. "
        "그래서 신청이력에서는 '접수 학기'와 '모집 차수'로 나눠 두고 원본 표기는 그대로 옆에 남겼으며, "
        "유학이력의 시작일에는 실제 학기 시작일(3월 1일 또는 9월 1일)을 넣었습니다.",
    )
    line(
        "학기 펼치는 법",
        "신청 건의 시작일과 종료일 사이에 걸치는 학기를 모두 만들었습니다. "
        "그 결과가 원본 '유학 학년도' 칸과 한 건도 어긋나지 않았습니다.",
    )
    line(
        "인원 세는 두 기준",
        "① 유학생 수(공식 기준) — 그 학기에 한 번이라도 유학한 인원. 도교육청 발표 자료와 같은 기준입니다. "
        "② 학기말 재적 — 그 학기 끝까지 남은 인원. 다음 학기로 넘어가는 인원과 같습니다. "
        "학기 중에 그만둔 학생만큼 ①이 ②보다 큽니다.",
    )
    line(
        "대조 결과",
        "발표 자료 '연도별 시·군 유학생 수'와 맞춰 본 결과, 2022학년도 27명·2023학년도 85명·"
        "2024학년도 165명·2025학년도 269명이 13개 시군 숫자까지 모두 일치했습니다. "
        "2026학년도 1학기는 발표 자료 333명, 계산 332명으로 고창에서 1명 차이가 나는데 "
        "학교 오보고분입니다. 이 332명 중 학기말까지 남는 인원은 325명입니다.",
    )
    line(
        "2026학년도 2학기",
        "확정 수치가 아닙니다. 들어 있는 인원은 최종배정 단계로 9월 1일 전학 전까지 포기할 수 있고, "
        "2차 모집 결과는 아직 원본에 입력되지 않았습니다. 계속 유학하는 학생도 연장이 확정된 뒤에 "
        "유학이력에 줄을 더해야 잡힙니다. 유학이력에서는 학기 상태가 '예정'으로 표시됩니다.",
    )
    line(
        "배정 판정",
        "'최종배정' 칸을 먼저 보고, 비어 있으면 종료일·배정희망서 칸을 차례로 봅니다. "
        "그래도 판단이 안 서는 건은 '확인필요'로 두고 데이터검증 시트에 모았습니다.",
    )
    line(
        "체제비 3년",
        "도청 지원은 최대 3년, 즉 6학기입니다. 시작 학기부터 여섯 번째 학기의 마지막 날을 만료일로 잡았습니다. "
        "원본 메모의 '2023년 3월 시작 → 2026년 2월까지' 와 같은 계산입니다.",
    )
    gap()

    head("주의")
    line(
        "개인정보",
        "학생·보호자 이름과 연락처가 그대로 들어 있습니다. 외부 공유용으로는 학생현황의 "
        "이름·연락처 열을 지운 사본을 따로 만들어 쓰십시오.",
    )
    line("체제비 금액", "설정 시트의 값은 2025. 전북 농촌유학 시행 지침 기준입니다. 지침이 바뀌면 그 시트를 고쳐야 합니다.")


# ─────────────────────────────────────────────────────────── 설정

RATE_CELLS = {
    "기준학기": "C5",
    "기준일": "C6",
    "가족1": "C9",
    "가족2": "C10",
    "가족3": "C11",
    "홈스테이": "C12",
    "센터": "C13",
    "도청": "C16",
    "지원학기": "C17",
    "서울기준학년도": "C20",
    "서울기준학기구분": "C21",
    "서울지원학기": "C22",
}


def sheet_settings(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 64

    t = ws.cell(2, 2, "설정")
    t.font = Font(name=FONT, size=16, bold=True, color=INK)
    n = ws.cell(3, 2, "노란 칸만 고치면 학생현황·체제비관리 시트가 다시 계산됩니다.")
    n.font = Font(name=FONT, size=10, color="595959")

    def band(row, text, unit="월 지원액(원)"):
        for col, val in ((2, text), (3, unit), (4, "근거")):
            c = ws.cell(row, col, val)
            c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=ACCENT)
            c.alignment = Alignment(
                horizontal="center" if col == 3 else "left", vertical="center", indent=1
            )
        ws.row_dimensions[row].height = 20

    def item(row, label, value, note, fmt="#,##0"):
        a = ws.cell(row, 2, label)
        a.font = Font(name=FONT, size=10)
        a.border = BOX
        a.alignment = Alignment(indent=1, vertical="center")
        b = ws.cell(row, 3, value)
        b.font = Font(name=FONT, size=10, bold=True, color="0000FF")
        b.fill = INPUT_FILL
        b.border = BOX
        b.number_format = fmt
        b.alignment = Alignment(horizontal="center", vertical="center")
        c = ws.cell(row, 4, note)
        c.font = Font(name=FONT, size=9, color="595959")
        c.border = BOX
        c.alignment = Alignment(indent=1, vertical="center", wrap_text=True)

    band(4, "기준 시점", "값")
    item(5, "기준 학기", T.sem_label(T.CURRENT_YEAR, T.CURRENT_TERM),
         "원본 종료일 칸의 '현재' 가 가리키는 학기", fmt="General")
    item(6, "기준일", dt.date.today(),
         "학생현황의 '현재 단계' 를 이 날짜로 판단합니다. "
         "유학 시작일이 이 날짜보다 뒤면 아직 '최종배정' 으로 봅니다.", fmt=DATE_FMT)

    src_note = "2025. 전북 농촌유학 시행 지침 Ⅳ. 사업 내용 — 유학경비 지원"
    band(8, "전북특별자치도교육청")
    item(9, "가족체류형 (가구 내 유학생 1명)", 300000, src_note)
    item(10, "가족체류형 (가구 내 유학생 2명)", 400000, src_note)
    item(11, "가족체류형 (가구 내 유학생 3명 이상)", 500000, src_note)
    item(12, "홈스테이형 (학생 1명당)", 300000, src_note)
    item(13, "유학센터형 (학생 1명당)", 300000, src_note)

    band(15, "전북특별자치도청")
    item(16, "도청 지원 (학생 1명당, 도10·시군10)", 200000, src_note + " / 최대 3년 지원")
    item(17, "도청 지원 기간 (학기)", 6,
         "최대 3년 = 6학기. 원본 메모: 2023년 3월 시작 → 2026년 2월까지 지원", fmt="0")

    band(19, "서울시교육청 (서울 원적 학생)", "값")
    item(20, "제출 기준 학년도", 2026, "서울원적_제출용 시트가 이 학기 명단으로 만들어집니다.", fmt="0")
    item(21, "제출 기준 학기 (1 또는 2)", 2, "1학기면 1, 2학기면 2", fmt="0")
    item(22, "지원 기간 (학기)", 6,
         "최초 참여 학기부터 세어 이 학기 수 이내면 '○'. "
         "서식 예시 3행(2022.1학기 ×, 2023.2학기 ×, 2024.2학기 ○)이 모두 6학기 경계로 갈립니다.",
         fmt="0")

    w = ws.cell(24, 2, "※ 유치원생은 다자녀 수에는 포함되나 유학경비 지원 대상에는 포함되지 않습니다(지침).")
    w.font = Font(name=FONT, size=9, color="A61C1C")


FAMILY_RANGE = "'설정'!$C$9:$C$11"


def rate(key: str) -> str:
    return f"'설정'!${RATE_CELLS[key][0]}${RATE_CELLS[key][1:]}"


# ─────────────────────────────────────────────────────────── 데이터 시트

# 1행 = 학생. '현재 단계' 는 옆 칸들을 보고 수식이 정한다.
STUDENT_HEADERS = [
    "학생ID", "성명", "성별", "현재 단계",
    "접수", "모집 차수", "가배정", "참가신청서", "최종배정",
    "유학 시작", "유학 종료", "종료 유형", "마지막 학기 마지막날",
    "현재 학년", "유학 지역", "유학 학교", "거주 유형", "거주지",
    "유학 학기수", "유학 학년도", "도청 지원 만료", "잔여 지원 학기",
    "최근 신청 결과", "미선정 단계", "미선정 사유", "신청 횟수", "단계 기록",
    "원 지역", "원 소속청", "원 소속교", "보호자 성명", "보호자 연락처",
    "가구ID", "가구 학생수", "학생 전화", "비고",
]

ENROLL_HEADERS = [
    "이력ID", "학기", "학년도", "학기구분", "학기순번", "학생ID", "성명", "구분",
    "유학 지역", "유학 학교", "학년", "성별", "거주 유형", "거주지",
    "가구ID", "원 지역", "원 소속교", "학기 상태", "학기말 재적", "종료 사유",
    "시작일", "종료일", "학기 마지막날", "신청ID",
]

APP_HEADERS = [
    "신청ID", "원본 행", "학생ID", "성명", "접수 학기", "접수 학년도", "학기구분", "모집 차수",
    "도달 단계", "배정 판정", "판정 근거", "미선정 단계", "미선정 사유(원문)", "미선정 사유(분류)",
    "접수 표기(원본)", "순(원본)", "예비유학생", "중학교 진학", "관내 전학", "기타",
    "유학 학년도(원본)", "배정 희망서", "참가 신청서", "최종 배정", "중간 종료 사유",
    "종료일(원본)", "학생 전화", "성별", "원 지역", "원 소속청", "원 소속교",
    "유학 지역", "전입시 학년", "현재 학년", "유학 학교", "이전 유학지역",
    "이전 유학학교", "거주 유형", "거주지", "보호자 성명", "보호자 연락처",
    "형제자매", "비고",
]


def sheet_students(ws, data):
    by_student = data["by_student"]
    prof = data["profiles"]
    rows = []
    for st in data["students"]:
        p = prof[st.student_id]
        lst = by_student.get(st.student_id, [])
        expire = None
        if lst:
            first = lst[0]
            expire = T.sem_end(*T.sem_from_index(T.sem_index(first.year, first.term) + 5))
        rows.append([
            st.student_id, st.name, st.gender,
            None,                                   # 현재 단계 (수식)
            p["접수"], p["모집 차수"], p["가배정"], p["참가신청서"], p["최종배정"],
            p["유학 시작"], p["유학 종료"],
            None,                                   # 종료 유형 (수식)
            p["마지막 학기 마지막날"],
            p["현재 학년"], p["최근 유학지역"], p["최근 유학학교"],
            p["최근 거주유형"], p["최근 거주지"],
            None,                                   # 유학 학기수 (수식)
            p["유학 학년도"], expire,
            None,                                   # 잔여 지원 학기 (수식)
            p["최근 신청 결과"], p["미선정 단계"], p["미선정 사유"],
            None,                                   # 신청 횟수 (수식)
            p["단계 기록 여부"],
            st.home_region, st.home_office, st.home_school,
            st.guardian, st.guardian_phone, st.household_id,
            None,                                   # 가구 학생수 (수식)
            st.phone, st.note,
        ])

    widths = {h: 12 for h in STUDENT_HEADERS}
    widths.update({
        "학생ID": 9, "성명": 11, "성별": 6, "현재 단계": 15,
        "접수": 11, "모집 차수": 8, "가배정": 11, "참가신청서": 11, "최종배정": 11,
        "유학 시작": 12, "유학 종료": 12, "종료 유형": 10, "마지막 학기 마지막날": 14,
        "현재 학년": 8, "유학 지역": 10, "유학 학교": 11, "거주 유형": 11, "거주지": 20,
        "유학 학기수": 8, "유학 학년도": 18, "도청 지원 만료": 13, "잔여 지원 학기": 10,
        "최근 신청 결과": 11, "미선정 단계": 10, "미선정 사유": 18, "신청 횟수": 8,
        "단계 기록": 9, "원 소속교": 14, "보호자 연락처": 15, "가구ID": 9,
        "가구 학생수": 8, "학생 전화": 14, "비고": 26,
    })
    last_row = write_table(
        ws, STUDENT_HEADERS, rows, widths,
        date_cols={"유학 시작", "유학 종료", "마지막 학기 마지막날", "도청 지원 만료"},
        wrap_cols={"비고", "미선정 사유", "거주지"},
        freeze="E",
    )

    C = {h: col_of(STUDENT_HEADERS, h) for h in STUDENT_HEADERS}
    e_student = rng("유학이력", ENROLL_HEADERS, "학생ID", LIMIT_ENROLL)
    a_student = rng("신청이력", APP_HEADERS, "학생ID", LIMIT_APP)

    for r in range(2, last_row + 1):
        # 끝난 학기에 맞춰 끝났으면 유학종료, 학기 도중이면 중도복귀
        put(ws, r, STUDENT_HEADERS.index("종료 유형") + 1,
            f'=IF(${C["유학 종료"]}{r}="","",'
            f'IF(${C["유학 종료"]}{r}>=${C["마지막 학기 마지막날"]}{r},"유학종료","중도복귀"))',
            fill=DERIVED_FILL)
        # 뒤 단계부터 차례로 본다. 아직 시작일이 오지 않았으면 최종배정에 머문다.
        put(ws, r, STUDENT_HEADERS.index("현재 단계") + 1,
            f'=IF(${C["유학 종료"]}{r}<>"",${C["종료 유형"]}{r},'
            f'IF(AND(${C["유학 시작"]}{r}<>"",${C["유학 시작"]}{r}<={rate("기준일")}),"유학중",'
            f'IF(${C["최종배정"]}{r}<>"","최종배정",'
            f'IF(${C["참가신청서"]}{r}<>"","참가신청서 제출",'
            f'IF(${C["가배정"]}{r}<>"","가배정","미선정·미전학")))))',
            bold=True, fill=DERIVED_FILL, align="center")
        put(ws, r, STUDENT_HEADERS.index("유학 학기수") + 1,
            f'=COUNTIF({e_student},${C["학생ID"]}{r})', fill=DERIVED_FILL, fmt="#,##0")
        put(ws, r, STUDENT_HEADERS.index("잔여 지원 학기") + 1,
            f'=IF(${C["유학 학기수"]}{r}=0,"",'
            f'MAX(0,{rate("지원학기")}-${C["유학 학기수"]}{r}))', fill=DERIVED_FILL, fmt="#,##0")
        put(ws, r, STUDENT_HEADERS.index("신청 횟수") + 1,
            f'=COUNTIF({a_student},${C["학생ID"]}{r})', fill=DERIVED_FILL, fmt="#,##0")
        put(ws, r, STUDENT_HEADERS.index("가구 학생수") + 1,
            f'=COUNTIF(${C["가구ID"]}$2:${C["가구ID"]}${LIMIT_STUDENT},${C["가구ID"]}{r})',
            fill=DERIVED_FILL, fmt="#,##0")

    add_dropdown(ws, STUDENT_HEADERS, "성별", ["남", "여"], LIMIT_STUDENT)
    add_dropdown(ws, STUDENT_HEADERS, "거주 유형",
                 ["가족체류형", "홈스테이형", "유학센터형"], LIMIT_STUDENT)

    stage = C["현재 단계"]
    for value, fill, color in (
        ("유학중", "D8F0DC", "17652A"),
        ("최종배정", "DCE9F7", "1F4E79"),
        ("참가신청서 제출", "FFF2CC", "7F6000"),
        ("가배정", "FFF2CC", "7F6000"),
        ("유학종료", "F2F2F2", "595959"),
        ("중도복귀", "FCE4E4", "A61C1C"),
        ("미선정·미전학", "FFFFFF", "A6A6A6"),
    ):
        ws.conditional_formatting.add(
            f"{stage}2:{stage}{LIMIT_STUDENT}",
            CellIsRule(operator="equal", formula=[f'"{value}"'],
                       fill=PatternFill("solid", fgColor=fill),
                       font=Font(name=FONT, size=10, bold=True, color=color)),
        )
    left = C["잔여 지원 학기"]
    ws.conditional_formatting.add(
        f"{left}2:{left}{LIMIT_STUDENT}",
        CellIsRule(operator="lessThanOrEqual", formula=["2"],
                   fill=PatternFill("solid", fgColor="FFF2CC"),
                   font=Font(name=FONT, size=10, bold=True, color="7F6000")),
    )
    return last_row


def sheet_enrollments(ws, data):
    rows = []
    for i, e in enumerate(data["enrollments"], start=1):
        rows.append([
            f"E{i:04d}", T.sem_label(e.year, e.term), e.year, f"{e.term}학기",
            T.sem_index(e.year, e.term), e.student_id, e.name, e.kind,
            e.region, e.school, e.grade, e.gender, e.residence, e.place,
            e.household_id, e.home_region, e.home_school, e.status,
            None,  # 학기말 재적 (수식)
            e.end_reason, e.start_date, e.end_date, e.term_end, e.app_id,
        ])

    last_row = write_table(
        ws, ENROLL_HEADERS, rows,
        {"이력ID": 8, "학기": 11, "학년도": 8, "학기구분": 8, "학기순번": 8,
         "학생ID": 9, "성명": 11, "구분": 8, "유학 지역": 10, "유학 학교": 11,
         "학년": 7, "성별": 6, "거주 유형": 11, "거주지": 22, "가구ID": 9,
         "원 지역": 10, "원 소속교": 13, "학기 상태": 9, "학기말 재적": 9,
         "종료 사유": 24, "시작일": 12, "종료일": 12, "학기 마지막날": 12, "신청ID": 9},
        date_cols={"시작일", "종료일", "학기 마지막날"},
        wrap_cols={"거주지", "종료 사유"},
    )

    # 학기 말까지 남아 있었는지 — 공식 인원은 이 기준으로 센다
    endc = col_of(ENROLL_HEADERS, "종료일")
    lastc = col_of(ENROLL_HEADERS, "학기 마지막날")
    for r in range(2, last_row + 1):
        c = ws.cell(r, ENROLL_HEADERS.index("학기말 재적") + 1,
                    f'=IF(OR(${endc}{r}="",${endc}{r}>=${lastc}{r}),"O","X")')
        c.font = Font(name=FONT, size=10)
        c.fill = DERIVED_FILL
        c.border = BOX
        c.alignment = Alignment(horizontal="center", vertical="center")

    add_dropdown(ws, ENROLL_HEADERS, "구분", ["신규", "계속", "재유학"], LIMIT_ENROLL)
    add_dropdown(ws, ENROLL_HEADERS, "학기 상태", ["재학", "종료", "예정"], LIMIT_ENROLL)
    add_dropdown(ws, ENROLL_HEADERS, "거주 유형", ["가족체류형", "홈스테이형", "유학센터형"], LIMIT_ENROLL)
    add_dropdown(ws, ENROLL_HEADERS, "학기구분", ["1학기", "2학기"], LIMIT_ENROLL)

    c = col_of(ENROLL_HEADERS, "학기 상태")
    ws.conditional_formatting.add(
        f"{c}2:{c}{LIMIT_ENROLL}",
        CellIsRule(operator="equal", formula=['"재학"'],
                   fill=PatternFill("solid", fgColor="D8F0DC"), font=Font(name=FONT, size=10, color="17652A")),
    )
    ws.conditional_formatting.add(
        f"{c}2:{c}{LIMIT_ENROLL}",
        CellIsRule(operator="equal", formula=['"예정"'],
                   fill=PatternFill("solid", fgColor="FFF2CC"), font=Font(name=FONT, size=10, color="7F6000")),
    )
    d = col_of(ENROLL_HEADERS, "구분")
    ws.conditional_formatting.add(
        f"{d}2:{d}{LIMIT_ENROLL}",
        CellIsRule(operator="equal", formula=['"신규"'], font=Font(name=FONT, size=10, bold=True, color=ACCENT)),
    )
    return last_row


def sheet_applications(ws, data):
    rows = []
    for a in data["applications"]:
        raw = a.raw
        rows.append([
            a.app_id, a.row, a.student_id, T.s(raw["성명"]),
            T.sem_label(a.intake_year, a.intake_term) if a.intake_year else None,
            a.intake_year, f"{a.intake_term}학기" if a.intake_term else None,
            f"{a.intake_round}차" if a.intake_round else None,
            a.stage, a.decision, a.decision_basis, a.reject_stage, a.reject_reason, a.reject_class,
            T.s(raw["접수일"]),
            raw["순"], T.s(raw["예비유학생"]), T.s(raw["중학교진학"]), T.s(raw["관내전학"]),
            T.s(raw["기타"]), T.s(raw["유학학년도_원본"]), T.s(raw["배정희망서"]),
            T.s(raw["참가신청서"]), T.s(raw["최종배정"]), T.s(raw["중간종료사유"]),
            T.s(raw["종료일"]), T.s(raw["학생전화"]), T.s(raw["성별"]), T.s(raw["원지역"]),
            T.s(raw["원소속청"]), T.s(raw["원소속교"]), T.s(raw["유학지역"]),
            T.s(raw["전입학년"]), T.s(raw["현재학년"]), T.s(raw["유학학교"]),
            T.s(raw["이전유학지역"]), T.s(raw["이전유학학교"]), T.s(raw["거주유형"]),
            T.s(raw["거주지"]), T.s(raw["보호자성명"]), T.s(raw["보호자연락처"]),
            T.s(raw["형제자매"]), T.s(raw["비고"]),
        ])

    widths = {h: 12 for h in APP_HEADERS}
    widths.update({
        "신청ID": 9, "원본 행": 8, "학생ID": 9, "성명": 11, "접수 학기": 11, "접수 학년도": 9,
        "학기구분": 8, "모집 차수": 8, "도달 단계": 16, "배정 판정": 10, "판정 근거": 34,
        "미선정 단계": 10, "미선정 사유(원문)": 30, "미선정 사유(분류)": 16, "접수 표기(원본)": 14,
        "중학교 진학": 26, "관내 전학": 22, "기타": 24, "유학 학년도(원본)": 16,
        "배정 희망서": 20, "중간 종료 사유": 20, "종료일(원본)": 16, "거주 유형": 16,
        "거주지": 22, "비고": 24, "순(원본)": 8,
    })
    last_row = write_table(
        ws, APP_HEADERS, rows, widths,
        date_cols=set(),
        wrap_cols={"판정 근거", "미선정 사유(원문)", "중학교 진학", "관내 전학", "기타",
                   "배정 희망서", "중간 종료 사유", "종료일(원본)", "거주 유형", "거주지", "비고"},
    )

    add_dropdown(ws, APP_HEADERS, "배정 판정", ["배정", "미배정", "확인필요"], LIMIT_APP)
    c = col_of(APP_HEADERS, "배정 판정")
    ws.conditional_formatting.add(
        f"{c}2:{c}{LIMIT_APP}",
        CellIsRule(operator="equal", formula=['"배정"'],
                   fill=PatternFill("solid", fgColor="D8F0DC"), font=Font(name=FONT, size=10, color="17652A")),
    )
    ws.conditional_formatting.add(
        f"{c}2:{c}{LIMIT_APP}",
        CellIsRule(operator="equal", formula=['"확인필요"'],
                   fill=PatternFill("solid", fgColor="FCE4E4"), font=Font(name=FONT, size=10, bold=True, color="A61C1C")),
    )
    return last_row


# ─────────────────────────────────────────────────────────── 집계


def style_head(ws, row, col, text, *, width=None, fill=INK, size=10):
    c = ws.cell(row, col, text)
    c.font = Font(name=FONT, size=size, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BOX
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def put(ws, row, col, value, *, bold=False, fill=None, fmt=None, align="center", size=10, color="000000"):
    c = ws.cell(row, col, value)
    c.font = Font(name=FONT, size=size, bold=bold, color=color)
    c.border = BOX
    c.alignment = Alignment(horizontal=align, vertical="center", indent=1 if align == "left" else 0)
    if fill:
        c.fill = fill if isinstance(fill, PatternFill) else PatternFill("solid", fgColor=fill)
    if fmt:
        c.number_format = fmt
    return c


def sheet_summary(ws, data):
    ws.sheet_view.showGridLines = False
    sems = data["semesters"]
    E = lambda name: rng("유학이력", ENROLL_HEADERS, name, LIMIT_ENROLL)  # noqa: E731
    A = lambda name: rng("신청이력", APP_HEADERS, name, LIMIT_APP)        # noqa: E731

    n_region = len({e.region for e in data["enrollments"] if e.region})
    n_school = len({(e.region, e.school) for e in data["enrollments"] if e.school})
    region_last, school_last = 2 + n_region, 2 + n_school

    t = ws.cell(1, 1, "학년도·학기별 유학생 현황")
    t.font = Font(name=FONT, size=15, bold=True, color=INK)
    n = ws.cell(2, 1,
                "'유학생 수(공식 기준)'는 그 학기에 한 번이라도 유학한 인원입니다. "
                "도교육청 발표 자료의 시군별 유학생 수와 같은 기준이며, "
                "2022~2025학년도는 13개 시군 숫자까지 모두 일치합니다. "
                "'학기말 재적'은 그 학기 끝까지 남은 인원으로, 다음 학기로 넘어가는 인원과 같습니다 "
                "(2026-1학기: 공식 332명 → 학기말 325명).")
    n.font = Font(name=FONT, size=9, color="595959")

    cols = ["학기", "유학생 수(공식 기준)", "학기말 재적", "학기 중 종료", "신규", "계속·재유학",
            "남", "여", "가족체류형", "홈스테이형", "유학센터형",
            "유학생 있는 지역 수", "유학생 있는 학교 수",
            "접수 건수", "배정", "미배정·확인필요"]
    r0 = 4
    for i, h in enumerate(cols, start=1):
        style_head(ws, r0, i, h, width=13 if i > 1 else 12)
    ws.row_dimensions[r0].height = 32

    for j, (y, term) in enumerate(sems):
        r = r0 + 1 + j
        lab = T.sem_label(y, term)
        q = f'"{lab}"'
        put(ws, r, 1, lab, bold=True, fill=SUB_FILL, align="center")
        put(ws, r, 2, f"=COUNTIF({E('학기')},{q})", bold=True, fmt="#,##0")
        put(ws, r, 3, f'=COUNTIFS({E("학기")},{q},{E("학기말 재적")},"O")', fmt="#,##0")
        put(ws, r, 4, f'=COUNTIFS({E("학기")},{q},{E("학기말 재적")},"X")', fmt="#,##0")
        put(ws, r, 5, f'=COUNTIFS({E("학기")},{q},{E("구분")},"신규")', fmt="#,##0")
        put(ws, r, 6, f'=COUNTIFS({E("학기")},{q},{E("구분")},"<>신규")', fmt="#,##0")
        put(ws, r, 7, f'=COUNTIFS({E("학기")},{q},{E("성별")},"남")', fmt="#,##0")
        put(ws, r, 8, f'=COUNTIFS({E("학기")},{q},{E("성별")},"여")', fmt="#,##0")
        put(ws, r, 9, f'=COUNTIFS({E("학기")},{q},{E("거주 유형")},"가족체류형")', fmt="#,##0")
        put(ws, r, 10, f'=COUNTIFS({E("학기")},{q},{E("거주 유형")},"홈스테이형")', fmt="#,##0")
        put(ws, r, 11, f'=COUNTIFS({E("학기")},{q},{E("거주 유형")},"유학센터형")', fmt="#,##0")
        put(ws, r, 12,
            f'=COUNTIF(지역별현황!{get_column_letter(2 + j)}$3:'
            f'{get_column_letter(2 + j)}${region_last},">0")', fmt="#,##0")
        put(ws, r, 13,
            f'=COUNTIF(학교별현황!{get_column_letter(3 + j)}$3:'
            f'{get_column_letter(3 + j)}${school_last},">0")', fmt="#,##0")
        put(ws, r, 14, f"=COUNTIF({A('접수 학기')},{q})", fmt="#,##0")
        put(ws, r, 15, f'=COUNTIFS({A("접수 학기")},{q},{A("배정 판정")},"배정")', fmt="#,##0")
        put(ws, r, 16, f'=COUNTIFS({A("접수 학기")},{q},{A("배정 판정")},"<>배정")', fmt="#,##0")

    last = r0 + len(sems)

    warn = ws.cell(last + 1, 1,
                   f"※ {T.sem_label(2026, 2)}는 확정 수치가 아닙니다. "
                   "① 표에 든 인원은 최종배정 단계로, 9월 1일 전학 전까지 참가를 포기할 수 있습니다. "
                   "② 2차 모집이 진행 중이며 그 결과는 아직 원본에 입력되지 않았습니다. "
                   "③ 계속 유학하는 학생은 연장이 확정된 뒤 유학이력에 줄을 더해야 잡힙니다 "
                   "— 2026-1학기 학기말 재적 325명이 그 대상입니다.")
    warn.font = Font(name=FONT, size=9, bold=True, color="A61C1C")

    # 학년도 요약: 1학기 인원 + 2학기에 새로 들어온 인원
    r1 = last + 3
    t2 = ws.cell(r1, 1, "학년도별 유학생 수 (연인원 아님 — 그 학년도에 한 번이라도 유학한 학생 수)")
    t2.font = Font(name=FONT, size=12, bold=True, color=INK)
    years = sorted({y for y, _ in sems})
    hdr = ["학년도", "유학생 수", "남", "여", "1학기", "2학기", "2학기 신규", "비고"]
    for i, h in enumerate(hdr, start=1):
        style_head(ws, r1 + 1, i, h, fill=ACCENT)
    notes = {
        2022: "발표 자료 27명과 일치 (시범사업 기간, 도청 체제비 미지원)",
        2023: "발표 자료 85명과 일치",
        2024: "발표 자료 165명과 일치",
        2025: "발표 자료 269명과 일치",
        2026: "발표 자료는 1학기 모집 기준 333명 — 고창 1명은 학교 오보고분이며 실제 332명",
    }
    for j, y in enumerate(years):
        r = r1 + 2 + j
        q1, q2 = f'"{T.sem_label(y, 1)}"', f'"{T.sem_label(y, 2)}"'
        put(ws, r, 1, f"{y}학년도", bold=True, fill=SUB_FILL)
        put(ws, r, 2,
            f"=COUNTIF({E('학기')},{q1})+COUNTIFS({E('학기')},{q2},{E('구분')},\"<>계속\")",
            bold=True, fmt="#,##0")
        for k, g in enumerate(("남", "여")):
            put(ws, r, 3 + k,
                f'=COUNTIFS({E("학기")},{q1},{E("성별")},"{g}")'
                f'+COUNTIFS({E("학기")},{q2},{E("구분")},"<>계속",{E("성별")},"{g}")',
                fmt="#,##0")
        put(ws, r, 5, f"=COUNTIF({E('학기')},{q1})", fmt="#,##0")
        put(ws, r, 6, f"=COUNTIF({E('학기')},{q2})", fmt="#,##0")
        put(ws, r, 7, f'=COUNTIFS({E("학기")},{q2},{E("구분")},"<>계속")', fmt="#,##0")
        put(ws, r, 8, notes.get(y, ""), align="left", size=9, color="595959")
    ws.column_dimensions["H"].width = 52

    # 원본에 적혀 있던 학기별 공식 수치가 어느 기준인지
    rb = r1 + 2 + len(years) + 1
    tb = ws.cell(rb, 1, "원본에 적혀 있던 공식 수치는 어느 기준인가")
    tb.font = Font(name=FONT, size=12, bold=True, color=INK)
    nb = ws.cell(rb + 1, 1,
                 "숨김 시트 '각종 현황'과 2025 시행 지침에 적힌 학기별 수치를 이 워크북과 맞춰 본 결과입니다. "
                 "학기마다 세는 기준이 달라서 그대로 이어 쓰면 어긋납니다.")
    nb.font = Font(name=FONT, size=9, color="595959")
    for i, h in enumerate(["구간", "원본 수치", "출처", "이 워크북", "읽는 법"], start=1):
        style_head(ws, rb + 2, i, h, fill=ACCENT, width=(14, 10, 20, 22, 64)[i - 1])
    checks = [
        ("2024학년도 1학기", 127, "각종 현황 시트",
         "연인원 132 · 학기말 재적 127",
         "학기말 재적과 같습니다. 연인원과는 5명 다릅니다."),
        ("2024학년도 2학기", 165, "각종 현황 시트",
         "연인원 160 · 학기말 재적 154",
         "어느 쪽과도 다릅니다. 165는 2024학년도 누계(1학기+2학기 신규)와 정확히 같고 "
         "발표 자료의 2024학년도 시군 숫자와도 맞아, 학년도 값이 2학기 칸에 적힌 것으로 보입니다."),
        ("2024. 9. 1. 기준", 163, "2025 시행 지침",
         "2024학년도 2학기 연인원 160",
         "시군마다 ±1~2씩 흩어져 어긋납니다. 원본 메모도 '중간집계일 기준'이라고 적고 있습니다."),
        ("2025학년도 1학기", 204, "각종 현황 시트",
         "연인원 204 · 학기말 재적 193",
         "연인원과 같습니다."),
        ("2025학년도 2학기", 257, "각종 현황 시트",
         "연인원 257 · 학기말 재적 241",
         "연인원과 같습니다."),
        ("서울 신규 2025. 2학기", 34, "각종 현황 시트",
         "배정 15 · 접수 33",
         "접수 기준에 가깝습니다. 2022~2025학년도 1학기까지는 서울 지원자가 모두 배정되어 "
         "두 기준이 같았는데, 2025학년도 2학기부터 미배정이 생기며 갈라졌습니다."),
    ]
    for j, (label, official, src, mine, how) in enumerate(checks):
        r = rb + 3 + j
        put(ws, r, 1, label, bold=True, fill=SUB_FILL, align="left")
        put(ws, r, 2, official, bold=True, fmt="#,##0")
        put(ws, r, 3, src, align="left", size=9, color="595959")
        put(ws, r, 4, mine, align="left", size=9)
        put(ws, r, 5, how, align="left", size=9, color="595959")
        ws.row_dimensions[r].height = 15 + 13 * (len(how) // 60)

    # 미선정 사유
    r2 = rb + 3 + len(checks) + 1
    t3 = ws.cell(r2, 1, "미선정 사유 (신청 건 기준)")
    t3.font = Font(name=FONT, size=12, bold=True, color=INK)
    reasons = sorted({a.reject_class for a in data["applications"] if a.reject_class})
    hdr2 = ["사유"] + [T.sem_label(y, t) for y, t in sems] + ["합계"]
    for i, h in enumerate(hdr2, start=1):
        style_head(ws, r2 + 1, i, h, fill=ACCENT)
    rc = col_of(APP_HEADERS, "미선정 사유(분류)")
    for j, reason in enumerate(reasons):
        r = r2 + 2 + j
        put(ws, r, 1, reason, align="left", fill=SUB_FILL)
        for k, (y, term) in enumerate(sems):
            put(ws, r, 2 + k,
                f'=COUNTIFS({A("접수 학기")},"{T.sem_label(y, term)}",'
                f"'신청이력'!${rc}$2:${rc}${LIMIT_APP},\"{reason}\")", fmt="#,##0")
        letters = f"{get_column_letter(2)}{r}:{get_column_letter(1 + len(sems))}{r}"
        put(ws, r, 2 + len(sems), f"=SUM({letters})", bold=True, fmt="#,##0")

    ws.freeze_panes = "B5"
    ws.column_dimensions["A"].width = 14


FUNNEL = [
    ("접수", None),
    ("가배정", "선정"),
    ("참가신청서 제출", "제출"),
    ("최종배정", None),
    ("실제 유학 시작", None),
]


def sheet_funnel(ws, data):
    """모집이 단계마다 얼마나 줄어드는지."""
    ws.sheet_view.showGridLines = False
    sems = data["semesters"]
    A = lambda name: rng("신청이력", APP_HEADERS, name, LIMIT_APP)   # noqa: E731
    E = lambda name: rng("유학이력", ENROLL_HEADERS, name, LIMIT_ENROLL)  # noqa: E731

    t = ws.cell(1, 1, "모집 단계별 현황")
    t.font = Font(name=FONT, size=15, bold=True, color=INK)
    n = ws.cell(2, 1,
                "접수한 학생이 단계마다 얼마나 남는지 봅니다. "
                f"학교 배정희망서와 학부모 참가신청서 칸은 "
                f"{T.sem_label(*T.STAGE_RECORDED_FROM)} 모집부터 기록되기 시작했습니다. "
                "그 전 학기는 결과만 남아 있어 가운데 두 단계가 비어 있습니다.")
    n.font = Font(name=FONT, size=9, color="595959")

    cols = ["학기", "접수", "가배정", "참가신청서 제출", "최종배정", "실제 유학 시작",
            "미선정·미전학", "확인 필요", "최종배정률", "단계 기록"]
    r0 = 4
    for i, h in enumerate(cols, start=1):
        style_head(ws, r0, i, h, width=16 if i == 1 else 14)
    ws.row_dimensions[r0].height = 32

    recorded = T.sem_index(*T.STAGE_RECORDED_FROM)
    wish = col_of(APP_HEADERS, "배정 희망서")
    submit = col_of(APP_HEADERS, "참가 신청서")
    for j, (y, term) in enumerate(sems):
        r = r0 + 1 + j
        lab = T.sem_label(y, term)
        q = f'"{lab}"'
        on = T.sem_index(y, term) >= recorded
        put(ws, r, 1, lab, bold=True, fill=SUB_FILL)
        put(ws, r, 2, f"=COUNTIF({A('접수 학기')},{q})", bold=True, fmt="#,##0")
        put(ws, r, 3,
            f'=COUNTIFS({A("접수 학기")},{q},'
            f"'신청이력'!${wish}$2:${wish}${LIMIT_APP},\"선정\")", fmt="#,##0")
        put(ws, r, 4,
            f'=COUNTIFS({A("접수 학기")},{q},'
            f"'신청이력'!${submit}$2:${submit}${LIMIT_APP},\"제출\")", fmt="#,##0")
        put(ws, r, 5, f'=COUNTIFS({A("접수 학기")},{q},{A("배정 판정")},"배정")',
            bold=True, fmt="#,##0")
        put(ws, r, 6, f'=COUNTIFS({E("학기")},{q},{E("구분")},"<>계속")', fmt="#,##0")
        put(ws, r, 7, f'=COUNTIFS({A("접수 학기")},{q},{A("배정 판정")},"미배정")', fmt="#,##0")
        put(ws, r, 8, f'=COUNTIFS({A("접수 학기")},{q},{A("배정 판정")},"확인필요")', fmt="#,##0")
        put(ws, r, 9, f"=IF($B{r}=0,\"\",$E{r}/$B{r})", fmt="0.0%")
        put(ws, r, 10, "기록" if on else "미기록", size=9,
            color="17652A" if on else "A6A6A6")

    r = r0 + 1 + len(sems)
    put(ws, r, 1, "합계", bold=True, fill=SUB_FILL)
    for c in range(2, 9):
        L = get_column_letter(c)
        put(ws, r, c, f"=SUM({L}{r0 + 1}:{L}{r - 1})", bold=True, fill=SUB_FILL, fmt="#,##0")
    put(ws, r, 9, f"=IF($B{r}=0,\"\",$E{r}/$B{r})", bold=True, fill=SUB_FILL, fmt="0.0%")
    put(ws, r, 10, "", fill=SUB_FILL)

    # 현재 학생들이 어느 단계에 있는지
    r2 = r + 2
    t2 = ws.cell(r2, 1, "학생 현재 단계 (1명 = 1행, 학생현황 시트 기준)")
    t2.font = Font(name=FONT, size=12, bold=True, color=INK)
    for i, h in enumerate(["단계", "인원", "설명"], start=1):
        style_head(ws, r2 + 1, i, h, fill=ACCENT, width=(16 if i == 1 else 14) if i < 3 else 62)
    notes = {
        "접수": "신청만 접수된 상태",
        "가배정": "학교가 배정 희망서에 '선정' 으로 올린 상태",
        "참가신청서 제출": "학부모가 최종 참가신청서까지 낸 상태",
        "최종배정": "도교육청 최종배정. 아직 전학 전이거나 시작일이 오지 않음",
        "유학중": "전학을 마치고 재학 중",
        "유학종료": "학기를 마치고 정상 종료",
        "중도복귀": "학기 도중에 그만두고 원적교로 복귀",
        "미선정·미전학": "어느 단계에서 빠졌거나 최종배정 후 전학하지 않음",
    }
    stage_col = col_of(STUDENT_HEADERS, "현재 단계")
    for j, name in enumerate(T.STAGES):
        rr = r2 + 2 + j
        put(ws, rr, 1, name, bold=True, fill=SUB_FILL, align="left")
        put(ws, rr, 2,
            f"=COUNTIF('학생현황'!${stage_col}$2:${stage_col}${LIMIT_STUDENT},\"{name}\")",
            bold=True, fmt="#,##0")
        put(ws, rr, 3, notes[name], align="left", size=9, color="595959")

    ws.freeze_panes = "B5"


def sheet_year(ws, data):
    """학년도 기준. 1학기와 2학기를 모두 다닌 학생도 한 번만 센다."""
    ws.sheet_view.showGridLines = False
    years = sorted({y for y, _ in data["semesters"]})
    E = lambda name: rng("유학이력", ENROLL_HEADERS, name, LIMIT_ENROLL)  # noqa: E731

    t = ws.cell(1, 1, "학년도별 현황")
    t.font = Font(name=FONT, size=15, bold=True, color=INK)
    n = ws.cell(2, 1,
                "학년도 인원은 '1학기 인원 + 2학기에 새로 온 인원'으로 셉니다. "
                "2학기의 '계속'은 이미 1학기에 세어졌기 때문입니다. 그래서 한 학생이 두 번 세어지지 않고, "
                "도교육청 발표 자료 '연도별 시·군 유학생 수'와 시군 숫자까지 같습니다.")
    n.font = Font(name=FONT, size=9, color="595959")

    def year_formula(year, extra_range=None, extra_key=None):
        q1, q2 = f'"{T.sem_label(year, 1)}"', f'"{T.sem_label(year, 2)}"'
        more = f",{extra_range},{extra_key}" if extra_range else ""
        return (f'=COUNTIFS({E("학기")},{q1}{more})'
                f'+COUNTIFS({E("학기")},{q2},{E("구분")},"<>계속"{more})')

    # ── 시군 × 학년도
    regions = sorted({e.region for e in data["enrollments"] if e.region})
    r0 = 4
    style_head(ws, r0, 1, "유학 지역", width=12)
    for k, y in enumerate(years):
        style_head(ws, r0, 2 + k, f"{y}학년도", width=12)
    style_head(ws, r0, 2 + len(years), "누적(연인원)", width=13)
    ws.row_dimensions[r0].height = 30
    for j, region in enumerate(regions):
        r = r0 + 1 + j
        put(ws, r, 1, region, bold=True, fill=SUB_FILL, align="left")
        for k, y in enumerate(years):
            put(ws, r, 2 + k, year_formula(y, E("유학 지역"), f"$A{r}"), fmt="#,##0")
        put(ws, r, 2 + len(years),
            f"=SUM(B{r}:{get_column_letter(1 + len(years))}{r})",
            bold=True, fmt="#,##0", fill=DERIVED_FILL)
    r = r0 + 1 + len(regions)
    put(ws, r, 1, "합계", bold=True, fill=SUB_FILL, align="left")
    for k in range(len(years) + 1):
        L = get_column_letter(2 + k)
        put(ws, r, 2 + k, f"=SUM({L}{r0 + 1}:{L}{r - 1})", bold=True, fmt="#,##0", fill=SUB_FILL)

    # ── 학교 × 학년도
    r2 = r + 3
    t2 = ws.cell(r2, 1, "학교별")
    t2.font = Font(name=FONT, size=12, bold=True, color=INK)
    schools = sorted({(e.region, e.school) for e in data["enrollments"] if e.school})
    r3 = r2 + 1
    style_head(ws, r3, 1, "유학 지역", fill=ACCENT)
    style_head(ws, r3, 2, "유학 학교", fill=ACCENT, width=14)
    for k, y in enumerate(years):
        style_head(ws, r3, 3 + k, f"{y}학년도", fill=ACCENT, width=12)
    style_head(ws, r3, 3 + len(years), "누적(연인원)", fill=ACCENT, width=13)
    for j, (region, school) in enumerate(schools):
        r = r3 + 1 + j
        put(ws, r, 1, region, fill=SUB_FILL, align="left")
        put(ws, r, 2, school, bold=True, align="left")
        for k, y in enumerate(years):
            put(ws, r, 3 + k, year_formula(y, E("유학 학교"), f"$B{r}"), fmt="#,##0")
        put(ws, r, 3 + len(years),
            f"=SUM(C{r}:{get_column_letter(2 + len(years))}{r})",
            bold=True, fmt="#,##0", fill=DERIVED_FILL)
    r = r3 + 1 + len(schools)
    put(ws, r, 1, "합계", bold=True, fill=SUB_FILL, align="left")
    put(ws, r, 2, "", fill=SUB_FILL)
    for k in range(len(years) + 1):
        L = get_column_letter(3 + k)
        put(ws, r, 3 + k, f"=SUM({L}{r3 + 1}:{L}{r - 1})", bold=True, fmt="#,##0", fill=SUB_FILL)

    ws.freeze_panes = "B5"


def sheet_seoul(ws, data):
    """서울 원적 학생만 따로. 서울시교육청 유학경비 지원 보고용.

    서울시교육청 양식을 아직 받지 못해 열 차례는 임시다. 양식이 오면 맞춘다.
    """
    prof = data["profiles"]
    by_student = data["by_student"]
    sems = [T.sem_label(y, t) for y, t in data["semesters"]]

    seoul = [st for st in data["students"]
             if any(e.home_region == "서울" for e in by_student.get(st.student_id, []))]

    headers = ["연번", "학생ID", "성명", "성별", "현재 학년", "원 소속청", "원 소속교",
               "유학 지역", "유학 학교", "거주 유형", "거주지",
               "유학 시작일", "유학 종료일", "누적 학기",
               "보호자 성명", "보호자 연락처", "학생 전화"] + sems + ["비고"]

    rows = []
    for i, st in enumerate(seoul, start=1):
        p = prof[st.student_id]
        rows.append([
            i, st.student_id, st.name, st.gender, p["현재 학년"],
            st.home_office, st.home_school,
            p["최근 유학지역"], p["최근 유학학교"], p["최근 거주유형"], p["최근 거주지"],
            p["유학 시작"], p["유학 종료"], None,
            st.guardian, st.guardian_phone, st.phone,
        ] + [None] * len(sems) + [st.note])

    widths = {h: 11 for h in headers}
    widths.update({
        "연번": 6, "학생ID": 9, "성명": 11, "성별": 6, "현재 학년": 8,
        "원 소속청": 12, "원 소속교": 14, "유학 지역": 10, "유학 학교": 11,
        "거주 유형": 11, "거주지": 20, "유학 시작일": 12, "유학 종료일": 12,
        "누적 학기": 8, "보호자 성명": 11, "보호자 연락처": 15, "학생 전화": 13, "비고": 24,
    })
    START = 8
    last_row = write_table(
        ws, headers, rows, widths,
        date_cols={"유학 시작일", "유학 종료일"},
        wrap_cols={"거주지", "비고"},
        start_row=START, freeze="D",
    )

    sid = col_of(headers, "학생ID")
    e_student = rng("유학이력", ENROLL_HEADERS, "학생ID", LIMIT_ENROLL)
    e_sem = rng("유학이력", ENROLL_HEADERS, "학기", LIMIT_ENROLL)
    for r in range(START + 1, last_row + 1):
        put(ws, r, headers.index("누적 학기") + 1,
            f"=COUNTIF({e_student},${sid}{r})", fill=DERIVED_FILL, fmt="#,##0")
        for k, lab in enumerate(sems):
            put(ws, r, headers.index(sems[0]) + 1 + k,
                f'=IF(COUNTIFS({e_student},${sid}{r},{e_sem},"{lab}")>0,"O","")',
                fill=DERIVED_FILL)

    # 머리말과 학기별 요약
    t = ws.cell(1, 1, "서울 원적 유학생")
    t.font = Font(name=FONT, size=15, bold=True, color=INK)
    n = ws.cell(2, 1,
                "서울시교육청 유학경비 지원 보고용입니다. 원 지역이 서울인 학생만 모았습니다. "
                "학기 칸의 'O' 는 그 학기에 재적했다는 뜻이라, 필요한 학기로 걸러 쓰면 됩니다. "
                "※ 서울시교육청 양식을 아직 받지 못해 열 차례는 임시입니다.")
    n.font = Font(name=FONT, size=9, color="595959")

    E = lambda name: rng("유학이력", ENROLL_HEADERS, name, LIMIT_ENROLL)  # noqa: E731
    style_head(ws, 4, 1, "학기", width=12)
    for k, lab in enumerate(sems):
        style_head(ws, 4, 2 + k, lab, width=12)
    put(ws, 5, 1, "서울 원적", bold=True, fill=SUB_FILL, align="left")
    put(ws, 6, 1, "전체", bold=True, fill=SUB_FILL, align="left")
    for k, lab in enumerate(sems):
        put(ws, 5, 2 + k, f'=COUNTIFS({E("학기")},"{lab}",{E("원 지역")},"서울")',
            bold=True, fmt="#,##0")
        put(ws, 6, 2 + k, f'=COUNTIF({E("학기")},"{lab}")', fmt="#,##0", color="595959")

    office = col_of(headers, "원 소속청")
    w = ws.cell(last_row + 2, 1,
                "※ 서울 지역교육지원청(원 소속청)이 비어 있는 학생이 있습니다. "
                "서울시교육청 보고에는 필요한 값이니 원본에서 채워 넣으십시오.")
    w.font = Font(name=FONT, size=9, color="A61C1C")
    put(ws, last_row + 3, 1, "원 소속청 미기재", bold=True, fill=SUB_FILL, align="left")
    put(ws, last_row + 3, 2,
        f'=COUNTBLANK(${office}${START + 1}:${office}${last_row})', bold=True, fmt="#,##0")
    return last_row


# 서울시교육청이 준 서식의 열 차례. 그대로 지킨다.
SEOUL_FORM = [
    "순", "유학지역\n(광역)", "유학지역\n(기초)", "유학학교", "학생 성명", "학년(2026)",
    "성별", "거주유형", "보호자 성명", "보호자 연락처", "원 교육지원청", "원 소속교",
    "최초 참여 기수", "비고", "서울시교육청 지원금 대상 여부",
]
# 서식에 없지만 검토에 필요한 칸. 제출할 때는 지운다.
SEOUL_EXTRA = ["학생ID", "최초 참여 학기순번", "기준 학기 기준 몇 번째", "명단에 오른 근거"]


def sheet_seoul_form(ws, data):
    """서울시교육청 제출 서식 그대로. 기준 학기에 참가하는 서울 원적 학생만."""
    prof = data["profiles"]
    by_student = data["by_student"]
    apps = {a.app_id: a for a in data["applications"]}
    base_y, base_t = 2026, 2                      # 설정 시트의 '제출 기준 학기'
    base_i = T.sem_index(base_y, base_t)
    prev_y, prev_t = T.sem_from_index(base_i - 1)

    rows = []
    for st in data["students"]:
        lst = by_student.get(st.student_id, [])
        if not lst or not any(e.home_region == "서울" for e in lst):
            continue
        here = next((e for e in lst if e.year == base_y and e.term == base_t), None)
        why = None
        if here:
            why = "기준 학기 배정"
        else:
            # 앞 학기에 재적하면서 종료일이 아직 '현재' 인 학생은 연장 후보다
            before = next((e for e in lst if e.year == prev_y and e.term == prev_t), None)
            if before:
                end = T.s(apps[before.app_id].raw["종료일"]) or ""
                if end.startswith("현재"):
                    here, why = before, "앞 학기 재적 · 연장 예정(원본 종료일 '현재')"
        if not here:
            continue
        first = lst[0]
        rows.append({
            "e": here, "st": st, "p": prof[st.student_id], "first": first, "why": why,
        })

    rows.sort(key=lambda r: ((r["e"].region or ""), (r["e"].school or ""), r["st"].name))

    # 서식 그대로: 2행 제목, 4행 머리글, 5행부터 데이터
    t = ws.cell(2, 2, f"(서식) {base_y}학년도 {base_t}학기 서울시교육청 농촌유학생 명단")
    t.font = Font(name=FONT, size=13, bold=True, color=INK)
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=13)

    headers = SEOUL_FORM + SEOUL_EXTRA
    for i, h in enumerate(headers, start=1):
        c = ws.cell(4, i, h)
        c.font = Font(name=FONT, size=10, bold=True,
                      color="FFFFFF" if i <= len(SEOUL_FORM) else "7F6000")
        c.fill = HEAD_FILL if i <= len(SEOUL_FORM) else INPUT_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
    ws.row_dimensions[4].height = 40

    for j, r in enumerate(rows):
        e, st, p, first = r["e"], r["st"], r["p"], r["first"]
        vals = [
            j + 1, "전북", e.region, e.school, st.name, e.grade, e.gender, e.residence,
            st.guardian, st.guardian_phone, st.home_office, st.home_school,
            f"{first.year}. {first.term}학기", ("형제자매" if T.s(apps[e.app_id].raw["형제자매"]) else None),
            None,                                   # 지원금 대상 여부 (수식)
            st.student_id, T.sem_index(first.year, first.term), None, r["why"],
        ]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(5 + j, i, v)
            c.font = Font(name=FONT, size=10,
                          color="000000" if i <= len(SEOUL_FORM) else "7F6000")
            c.border = BOX
            c.alignment = Alignment(vertical="center",
                                    horizontal="center" if i in (1, 6, 7, 15) else None,
                                    wrap_text=(i == len(headers)))
            if i > len(SEOUL_FORM):
                c.fill = PatternFill("solid", fgColor="FFF9E6")

    last = 4 + len(rows)
    P = {h: get_column_letter(i + 1) for i, h in enumerate(headers)}
    for j in range(len(rows)):
        r = 5 + j
        put(ws, r, headers.index("기준 학기 기준 몇 번째") + 1,
            f'=({rate("서울기준학년도")}*2+{rate("서울기준학기구분")}-1)'
            f'-${P["최초 참여 학기순번"]}{r}+1',
            fill="FFF9E6", fmt="0", color="7F6000")
        put(ws, r, headers.index("서울시교육청 지원금 대상 여부") + 1,
            f'=IF(${P["기준 학기 기준 몇 번째"]}{r}<={rate("서울지원학기")},"○","×")',
            bold=True)

    widths = {"순": 6, "유학지역\n(광역)": 10, "유학지역\n(기초)": 10, "유학학교": 12,
              "학생 성명": 11, "학년(2026)": 10, "성별": 6, "거주유형": 11,
              "보호자 성명": 11, "보호자 연락처": 15, "원 교육지원청": 13, "원 소속교": 14,
              "최초 참여 기수": 13, "비고": 12, "서울시교육청 지원금 대상 여부": 14,
              "학생ID": 9, "최초 참여 학기순번": 10, "기준 학기 기준 몇 번째": 10,
              "명단에 오른 근거": 30}
    for h, w in widths.items():
        ws.column_dimensions[P[h]].width = w
    ws.freeze_panes = "F5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{max(last, 5)}"

    n = ws.cell(last + 2, 2,
                f"※ 왼쪽 {len(SEOUL_FORM)}개 열이 서울시교육청 서식 그대로입니다. "
                f"노란 {len(SEOUL_EXTRA)}개 열은 검토용이니 제출할 때 지우십시오.")
    n.font = Font(name=FONT, size=9, color="595959")
    n2 = ws.cell(last + 3, 2,
                 "※ 지원금 대상 여부는 '최초 참여 학기부터 세어 6학기 이내면 ○' 로 계산했습니다. "
                 "서식 예시 3행이 모두 이 경계로 갈려 그렇게 읽었습니다. 기준이 다르면 설정 시트에서 고치십시오.")
    n2.font = Font(name=FONT, size=9, color="A61C1C")
    n3 = ws.cell(last + 4, 2,
                 f"※ 원본에 {base_y}학년도 {base_t}학기 연장이 아직 입력되지 않아, "
                 "앞 학기 재적자 중 종료일이 '현재' 인 학생을 연장 예정으로 함께 올렸습니다. "
                 "'명단에 오른 근거' 열에서 갈라 보고, 연장하지 않는 학생은 지우십시오.")
    n3.font = Font(name=FONT, size=9, color="A61C1C")

    office = P["원 교육지원청"]
    put(ws, last + 6, 2, "원 교육지원청이 비어 있는 학생", bold=True, fill=SUB_FILL, align="left")
    put(ws, last + 6, 3, f"=COUNTBLANK(${office}$5:${office}${max(last, 5)})",
        bold=True, fmt="#,##0")
    n4 = ws.cell(last + 7, 2, "서울시교육청 보고에 필요한 값입니다. 원본 '원 소속청' 칸을 채우십시오.")
    n4.font = Font(name=FONT, size=9, color="A61C1C")
    return len(rows)


def sheet_region(ws, data):
    ws.sheet_view.showGridLines = False
    sems = data["semesters"]
    regions = sorted({e.region for e in data["enrollments"] if e.region})
    E = lambda name: rng("유학이력", ENROLL_HEADERS, name, LIMIT_ENROLL)  # noqa: E731

    t = ws.cell(1, 1, "시군별 유학생 수")
    t.font = Font(name=FONT, size=15, bold=True, color=INK)
    n = ws.cell(1, 3,
                "학년도로 묶으면 도교육청 발표 자료('연도별 시·군 유학생 수')와 "
                "2022~2025학년도 13개 시군 숫자가 모두 일치합니다.")
    n.font = Font(name=FONT, size=9, color="595959")
    style_head(ws, 2, 1, "유학 지역", width=12)
    for k, (y, term) in enumerate(sems):
        style_head(ws, 2, 2 + k, T.sem_label(y, term), width=11)
    style_head(ws, 2, 2 + len(sems), "누적(연인원)", width=13)
    ws.row_dimensions[2].height = 30

    for j, region in enumerate(regions):
        r = 3 + j
        put(ws, r, 1, region, bold=True, fill=SUB_FILL, align="left")
        for k, (y, term) in enumerate(sems):
            put(ws, r, 2 + k,
                f'=COUNTIFS({E("학기")},"{T.sem_label(y, term)}",{E("유학 지역")},$A{r})',
                fmt="#,##0")
        put(ws, r, 2 + len(sems),
            f"=SUM(B{r}:{get_column_letter(1 + len(sems))}{r})", bold=True, fmt="#,##0", fill=DERIVED_FILL)

    r = 3 + len(regions)
    put(ws, r, 1, "합계", bold=True, fill=SUB_FILL, align="left")
    for k in range(len(sems) + 1):
        L = get_column_letter(2 + k)
        put(ws, r, 2 + k, f"=SUM({L}3:{L}{r - 1})", bold=True, fmt="#,##0", fill=SUB_FILL)
    ws.freeze_panes = "B3"


def sheet_school(ws, data):
    ws.sheet_view.showGridLines = False
    sems = data["semesters"]
    schools = sorted({(e.region, e.school) for e in data["enrollments"] if e.school})
    E = lambda name: rng("유학이력", ENROLL_HEADERS, name, LIMIT_ENROLL)  # noqa: E731

    t = ws.cell(1, 1, "운영학교별 유학생 수")
    t.font = Font(name=FONT, size=15, bold=True, color=INK)
    style_head(ws, 2, 1, "유학 지역", width=12)
    style_head(ws, 2, 2, "유학 학교", width=14)
    for k, (y, term) in enumerate(sems):
        style_head(ws, 2, 3 + k, T.sem_label(y, term), width=11)
    style_head(ws, 2, 3 + len(sems), "누적(연인원)", width=13)
    ws.row_dimensions[2].height = 30

    for j, (region, school) in enumerate(schools):
        r = 3 + j
        put(ws, r, 1, region, fill=SUB_FILL, align="left")
        put(ws, r, 2, school, bold=True, align="left")
        for k, (y, term) in enumerate(sems):
            put(ws, r, 3 + k,
                f'=COUNTIFS({E("학기")},"{T.sem_label(y, term)}",{E("유학 학교")},$B{r})',
                fmt="#,##0")
        put(ws, r, 3 + len(sems),
            f"=SUM(C{r}:{get_column_letter(2 + len(sems))}{r})", bold=True, fmt="#,##0", fill=DERIVED_FILL)

    r = 3 + len(schools)
    put(ws, r, 1, "합계", bold=True, fill=SUB_FILL, align="left")
    put(ws, r, 2, "", fill=SUB_FILL)
    for k in range(len(sems) + 1):
        L = get_column_letter(3 + k)
        put(ws, r, 3 + k, f"=SUM({L}3:{L}{r - 1})", bold=True, fmt="#,##0", fill=SUB_FILL)

    ws.auto_filter.ref = f"A2:{get_column_letter(3 + len(sems))}{r - 1}"
    ws.freeze_panes = "C3"


COST_HEADERS = [
    "학생ID", "성명", "가구ID", "가구 대표", "유학 상태", "기준 학기", "유학 지역", "유학 학교",
    "거주 유형", "지원 시작 학기", "지원 시작일", "도청 지원 만료일", "이수 학기",
    "잔여 지원 학기", "만료 알림", "월 체제비(교육청)", "월 체제비(도청)",
]


def sheet_cost(ws, data):
    by_student = data["by_student"]
    rows = []
    seen_house = set()
    for st in data["students"]:
        lst = by_student.get(st.student_id, [])
        if not lst:
            continue
        first, last = lst[0], lst[-1]
        start_i = T.sem_index(first.year, first.term)
        expire = T.sem_end(*T.sem_from_index(start_i + 5))
        rep = st.household_id not in seen_house
        seen_house.add(st.household_id)
        status = {"재학": "재학중", "예정": "배정예정", "종료": "종료"}[last.status]
        rows.append([
            st.student_id, st.name, st.household_id, "대표" if rep else "-", status,
            T.sem_label(last.year, last.term), last.region, last.school, last.residence,
            T.sem_label(first.year, first.term), first.start_date, expire,
            None, None, None, None, None,
        ])
    rows.sort(key=lambda x: ({"재학중": 0, "배정예정": 1, "종료": 2}[x[4]], x[11]))

    last_row = write_table(
        ws, COST_HEADERS, rows,
        {"학생ID": 9, "성명": 11, "가구ID": 9, "가구 대표": 9, "유학 상태": 10,
         "기준 학기": 11, "유학 지역": 10, "유학 학교": 11, "거주 유형": 11, "지원 시작 학기": 12,
         "지원 시작일": 12, "도청 지원 만료일": 14, "이수 학기": 9,
         "잔여 지원 학기": 11, "만료 알림": 12, "월 체제비(교육청)": 15, "월 체제비(도청)": 14},
        date_cols={"지원 시작일", "도청 지원 만료일"},
    )

    sid = col_of(COST_HEADERS, "학생ID")
    hid = col_of(COST_HEADERS, "가구ID")
    rep = col_of(COST_HEADERS, "가구 대표")
    stat = col_of(COST_HEADERS, "유학 상태")
    res = col_of(COST_HEADERS, "거주 유형")
    done = col_of(COST_HEADERS, "이수 학기")
    left = col_of(COST_HEADERS, "잔여 지원 학기")
    e_student = rng("유학이력", ENROLL_HEADERS, "학생ID", LIMIT_ENROLL)
    base = col_of(COST_HEADERS, "기준 학기")
    e_sem = rng("유학이력", ENROLL_HEADERS, "학기", LIMIT_ENROLL)

    for r in range(2, last_row + 1):
        put(ws, r, COST_HEADERS.index("이수 학기") + 1,
            f"=COUNTIF({e_student},${sid}{r})", fill="F2F2F2", fmt="#,##0")
        put(ws, r, COST_HEADERS.index("잔여 지원 학기") + 1,
            f"=MAX(0,{rate('지원학기')}-${done}{r})", fill="F2F2F2", fmt="#,##0")
        put(ws, r, COST_HEADERS.index("만료 알림") + 1,
            f'=IF(${stat}{r}="종료","-",'
            f'IF(${left}{r}=0,"지원 종료",'
            f'IF(${left}{r}<=2,"만료 임박","")))',
            fill="F2F2F2")
        # 가족체류형은 가구당 지급이라 가구 대표 행에만 금액을 둔다.
        put(ws, r, COST_HEADERS.index("월 체제비(교육청)") + 1,
            f'=IF(${stat}{r}="종료",0,'
            f'IF(${res}{r}="가족체류형",'
            f'IF(${rep}{r}<>"대표",0,'
            f"INDEX({FAMILY_RANGE},MIN(3,MAX(1,"
            f"COUNTIFS({rng('유학이력', ENROLL_HEADERS, '가구ID', LIMIT_ENROLL)},${hid}{r},"
            f"{e_sem},${base}{r}))))),"
            f'IF(${res}{r}="홈스테이형",{rate("홈스테이")},{rate("센터")})))',
            fill="F2F2F2", fmt="#,##0")
        put(ws, r, COST_HEADERS.index("월 체제비(도청)") + 1,
            f'=IF(OR(${stat}{r}="종료",${left}{r}=0),0,{rate("도청")})',
            fill="F2F2F2", fmt="#,##0")

    r = last_row + 1
    put(ws, r, 1, "합계", bold=True, fill=SUB_FILL, align="left")
    for i in range(2, len(COST_HEADERS) + 1):
        put(ws, r, i, "", fill=SUB_FILL)
    for name in ("월 체제비(교육청)", "월 체제비(도청)"):
        L = col_of(COST_HEADERS, name)
        put(ws, r, COST_HEADERS.index(name) + 1, f"=SUM({L}2:{L}{last_row})",
            bold=True, fill=SUB_FILL, fmt="#,##0")
    note = ws.cell(r + 2, 1,
                   "※ 월 체제비는 설정 시트의 단가로 계산한 참고값입니다. "
                   "가족체류형은 가구당 지급이라 '가구 대표' 행에만 금액이 잡힙니다. "
                   "가구 내 유학생 수는 그 학생의 '기준 학기'(가장 최근 유학 학기) 기준으로 셉니다.")
    note.font = Font(name=FONT, size=9, color="595959")

    c = col_of(COST_HEADERS, "만료 알림")
    ws.conditional_formatting.add(
        f"{c}2:{c}{last_row}",
        CellIsRule(operator="equal", formula=['"만료 임박"'],
                   fill=PatternFill("solid", fgColor="FFF2CC"), font=Font(name=FONT, size=10, bold=True, color="7F6000")),
    )
    ws.conditional_formatting.add(
        f"{c}2:{c}{last_row}",
        CellIsRule(operator="equal", formula=['"지원 종료"'],
                   fill=PatternFill("solid", fgColor="FCE4E4"), font=Font(name=FONT, size=10, bold=True, color="A61C1C")),
    )


ISSUE_HEADERS = ["번호", "구분", "집계 영향", "유형", "학생ID", "성명", "신청ID",
                 "원본 행", "내용", "처리 메모"]


# 도교육청 발표 자료 '연도별 시·군 유학생 수' ('26. 1학기 모집 기준)
OFFICIAL = [
    (2022, 27, "일치"), (2023, 85, "일치"), (2024, 165, "일치"), (2025, 269, "일치"),
    (2026, 333, "고창 1명 차이 — 학교 오보고분, 실제 332명"),
]

START = 9


def raw_text(v):
    """다듬기 전의 원본 값을 글자로."""
    if v is None:
        return ""
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def reconcile(data):
    """원본 값을 옮기면서 무엇을 다듬었는지, 원본 한 줄이 어디로 갔는지."""
    trimmed = []
    cells = 0
    for a in data["applications"]:
        for key in T.COL:
            cells += 1
            v = a.raw[key]
            before, after = raw_text(v), (T.s(v) or "")
            if before != after:
                trimmed.append((a.row, key, before, after))
    verdicts = Counter(a.decision for a in data["applications"])
    return {
        "rows": len(data["applications"]),
        "cells": cells,
        "trimmed": trimmed,
        "verdicts": verdicts,
        "enrollments": len(data["enrollments"]),
    }


def sheet_issues(ws, data):
    t = ws.cell(1, 1, "데이터 검증")
    t.font = Font(name=FONT, size=15, bold=True, color=INK)
    n = ws.cell(2, 1,
                "먼저 원본과 어긋난 곳이 없는지, 그 다음 발표 자료와 숫자가 맞는지 보고, "
                "마지막에 손볼 것을 추립니다. '집계 영향 없음'인 항목은 숫자를 흔들지 않으므로 "
                "급히 고치지 않아도 됩니다.")
    n.font = Font(name=FONT, size=9, color="595959")

    rec = reconcile(data)
    r = 3
    for i, h in enumerate(["원본과 맞춰 본 것", "결과", "설명"], start=1):
        style_head(ws, r, i, h, fill=ACCENT, width=(24 if i == 1 else 14) if i < 3 else 62)
    lines = [
        ("원본 데이터 행", f"{rec['rows']:,}행", "신청이력이 원본 한 줄에 한 줄씩 맞물립니다. '원본 행' 열로 찾아가세요."),
        ("옮긴 칸", f"{rec['cells']:,}칸",
         f"원본 30개 열을 모두 신청이력에 그대로 두었습니다. 그중 {len(rec['trimmed'])}칸만 다듬었습니다."),
        ("배정 → 유학이력", f"{rec['verdicts'].get('배정', 0):,}건",
         f"유학이력 {rec['enrollments']:,}행으로 펼쳤습니다"
         f"(한 건당 평균 {rec['enrollments'] / max(rec['verdicts'].get('배정', 1), 1):.1f}학기)."),
        ("미배정", f"{rec['verdicts'].get('미배정', 0):,}건", "신청이력에만 남습니다. 유학하지 않은 건입니다."),
        ("확인필요", f"{rec['verdicts'].get('확인필요', 0):,}건", "신청이력에 남기고 아래 목록에 올렸습니다."),
    ]
    for j, (a, b, c) in enumerate(lines):
        rr = r + 1 + j
        put(ws, rr, 1, a, bold=True, fill=SUB_FILL, align="left")
        put(ws, rr, 2, b, bold=True)
        put(ws, rr, 3, c, align="left", size=9, color="595959")
    r += 1 + len(lines)

    if rec["trimmed"]:
        put(ws, r + 1, 1, "다듬은 칸", bold=True, fill=SUB_FILL, align="left")
        put(ws, r + 1, 2, f"{len(rec['trimmed'])}칸", bold=True)
        put(ws, r + 1, 3,
            " · ".join(f"{row}행 {key}: '{before}' → '{after}'"
                       for row, key, before, after in rec["trimmed"][:6]),
            align="left", size=9, color="595959")
        r += 1
    r += 2

    for i, h in enumerate(["학년도", "발표 자료", "워크북", "시군별 대조"], start=1):
        style_head(ws, r, i, h, fill=ACCENT, width=(11 if i < 4 else 62))
    by_year = {}
    for e in data["enrollments"]:
        if e.year == 2026 and e.term != 1:
            continue
        by_year.setdefault(e.year, set()).add(e.student_id)
    for j, (year, official, note) in enumerate(OFFICIAL):
        rr = r + 1 + j
        put(ws, rr, 1, f"{year}학년도" + (" 1학기" if year == 2026 else ""), bold=True, fill=SUB_FILL)
        put(ws, rr, 2, official, fmt="#,##0")
        put(ws, rr, 3, len(by_year.get(year, ())), bold=True, fmt="#,##0")
        put(ws, rr, 4, note, align="left", size=9,
            color="A61C1C" if note != "일치" else "17652A")

    rows = [
        [i, x["심각도"], x["집계 영향"], x["유형"], x["학생ID"], x["성명"], x["신청ID"],
         x["원본행"], x["내용"], None]
        for i, x in enumerate(data["issues"], start=1)
    ]
    last_row = write_table(
        ws, ISSUE_HEADERS, rows,
        {"번호": 6, "구분": 10, "집계 영향": 9, "유형": 18, "학생ID": 14, "성명": 11,
         "신청ID": 14, "원본 행": 10, "내용": 74, "처리 메모": 22},
        wrap_cols={"내용", "처리 메모"},
        start_row=r + len(OFFICIAL) + 3,
    )
    head_row = last_row - len(data["issues"])
    add_dropdown(ws, ISSUE_HEADERS, "처리 메모", ["확인 완료", "수정함", "원본이 맞음", "보류"], last_row)
    c = col_of(ISSUE_HEADERS, "구분")
    ws.conditional_formatting.add(
        f"A{head_row + 1}:{get_column_letter(len(ISSUE_HEADERS))}{last_row}",
        FormulaRule(formula=[f'${c}{head_row + 1}="조치 필요"'],
                    fill=PatternFill("solid", fgColor="FCE4E4")),
    )
    ws.conditional_formatting.add(
        f"{c}{head_row + 1}:{c}{last_row}",
        CellIsRule(operator="equal", formula=['"참고"'],
                   font=Font(name=FONT, size=10, color="808080")),
    )
    for rr in range(head_row + 1, last_row + 1):
        ws.cell(rr, ISSUE_HEADERS.index("처리 메모") + 1).fill = INPUT_FILL


def sheet_source(ws, src_path):
    wb = openpyxl.load_workbook(src_path, data_only=True)
    src = wb[T.SOURCE_SHEET]
    for r in range(1, src.max_row + 1):
        for c in range(1, 31):
            v = src.cell(r, c).value
            if v is None:
                continue
            cell = ws.cell(r, c, v)
            cell.font = Font(name=FONT, size=9, bold=(r == T.HEADER_ROW))
            if isinstance(v, dt.datetime):
                cell.number_format = DATE_FMT
            if r == T.HEADER_ROW:
                cell.fill = SUB_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 31):
        ws.column_dimensions[get_column_letter(c)].width = 13
    ws.freeze_panes = "A9"
    ws.row_dimensions[T.HEADER_ROW].height = 40


def main(src, out):
    data = T.build(src)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sheet_guide(wb.create_sheet("안내"), data, src.split("/")[-1])
    sheet_students(wb.create_sheet("학생현황"), data)
    sheet_enrollments(wb.create_sheet("유학이력"), data)
    sheet_applications(wb.create_sheet("신청이력"), data)
    sheet_funnel(wb.create_sheet("모집단계현황"), data)
    sheet_summary(wb.create_sheet("학기별집계"), data)
    sheet_year(wb.create_sheet("연도별현황"), data)
    sheet_seoul(wb.create_sheet("서울원적"), data)
    n_form = sheet_seoul_form(wb.create_sheet("서울원적_제출용"), data)
    sheet_region(wb.create_sheet("지역별현황"), data)
    sheet_school(wb.create_sheet("학교별현황"), data)
    sheet_cost(wb.create_sheet("체제비관리"), data)
    sheet_settings(wb.create_sheet("설정"))
    sheet_issues(wb.create_sheet("데이터검증"), data)
    sheet_source(wb.create_sheet("원본_전체"), src)

    for ws in wb.worksheets:
        ws.sheet_properties.tabColor = ACCENT if ws.title in {"학생현황", "유학이력", "신청이력"} else None
    wb.save(out)
    print(f"저장: {out}")
    print(f"  학생 {len(data['students'])} / 신청 {len(data['applications'])} / "
          f"유학이력 {len(data['enrollments'])} / 검증 {len(data['issues'])}")


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
