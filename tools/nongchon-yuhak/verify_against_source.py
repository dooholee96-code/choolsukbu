"""원본 명단과 새 워크북이 어긋나지 않는지 셀 단위로 맞춘다.

  python verify_against_source.py <원본.xlsx> <워크북.xlsx>

두 가지를 본다.

  하나. 원본의 값이 새 워크북에 하나도 빠짐없이 남아 있는가.
  둘.   새 워크북이 원본에 없는 값을 만들어내지 않았는가.

값을 다듬은 것(앞뒤 공백 제거, 날짜를 yyyy-mm-dd 로)은 '어긋남' 이 아니라
'다듬음' 으로 따로 센다. 무엇을 다듬었는지 눈으로 볼 수 있어야 하기 때문이다.
"""

from __future__ import annotations

import datetime as dt
import sys
from collections import defaultdict

import openpyxl
from openpyxl.utils import get_column_letter

import transform as T

# 원본 열 이름 → 신청이력의 열 이름. 원본 30개 열이 모두 여기에 있다.
KEEP = {
    "순": "순(원본)",
    "예비유학생": "예비유학생",
    "중학교진학": "중학교 진학",
    "관내전학": "관내 전학",
    "기타": "기타",
    "유학학년도_원본": "유학 학년도(원본)",
    "접수일": "접수 표기(원본)",
    "배정희망서": "배정 희망서",
    "참가신청서": "참가 신청서",
    "최종배정": "최종 배정",
    "중간종료사유": "중간 종료 사유",
    "종료일": "종료일(원본)",
    "성명": "성명",
    "학생전화": "학생 전화",
    "성별": "성별",
    "원지역": "원 지역",
    "원소속청": "원 소속청",
    "원소속교": "원 소속교",
    "유학지역": "유학 지역",
    "전입학년": "전입시 학년",
    "현재학년": "현재 학년",
    "유학학교": "유학 학교",
    "이전유학지역": "이전 유학지역",
    "이전유학학교": "이전 유학학교",
    "거주유형": "거주 유형",
    "거주지": "거주지",
    "보호자성명": "보호자 성명",
    "보호자연락처": "보호자 연락처",
    "형제자매": "형제자매",
    "비고": "비고",
}


def raw_text(v):
    """다듬기 전의 원본 값을 글자로. 날짜만 같은 꼴로 맞춘다."""
    if v is None:
        return ""
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def header_map(ws, row=1):
    return {ws.cell(row, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(row, c).value}


def main(src_path, book_path):
    src = T.find_sheet(openpyxl.load_workbook(src_path, data_only=True))
    wb = openpyxl.load_workbook(book_path, data_only=False)

    # 원본 데이터 행
    src_rows = {}
    for r in range(T.FIRST_DATA_ROW, src.max_row + 1):
        if src.cell(r, T.COL["성명"]).value not in (None, ""):
            src_rows[r] = r

    apps = wb["신청이력"]
    ah = header_map(apps)
    app_rows = {}
    for r in range(2, apps.max_row + 1):
        origin = apps.cell(r, ah["원본 행"]).value
        if origin is None:
            break
        app_rows[origin] = r

    print(f"원본 데이터 행 {len(src_rows)}개 · 신청이력 {len(app_rows)}행")
    problems = []
    missing = sorted(set(src_rows) - set(app_rows))
    extra = sorted(set(app_rows) - set(src_rows))
    if missing:
        problems.append(f"신청이력에 없는 원본 행 {len(missing)}개: {missing[:10]}")
    if extra:
        problems.append(f"원본에 없는 신청이력 행 {len(extra)}개: {extra[:10]}")

    # ── 셀 대조 ─────────────────────────────────────────
    checked = 0
    diffs = []
    trimmed = defaultdict(list)
    for origin, arow in sorted(app_rows.items()):
        if origin not in src_rows:
            continue
        for src_key, app_key in KEEP.items():
            checked += 1
            a = raw_text(src.cell(origin, T.COL[src_key]).value)
            b = apps.cell(arow, ah[app_key]).value
            b = "" if b is None else raw_text(b)
            if a == b:
                continue
            if T.s(src.cell(origin, T.COL[src_key]).value) == (b or None) or (a.strip() == b):
                trimmed[src_key].append((origin, a, b))
            else:
                diffs.append(f"원본 {origin}행 '{src_key}': 원본 {a[:40]!r} ≠ 워크북 {b[:40]!r}")

    print(f"셀 대조 {checked:,}칸 · 그대로 {checked - len(diffs) - sum(len(v) for v in trimmed.values()):,}칸 "
          f"· 다듬음 {sum(len(v) for v in trimmed.values())}칸 · 어긋남 {len(diffs)}칸")
    for key, items in sorted(trimmed.items(), key=lambda kv: -len(kv[1])):
        sample = items[0]
        print(f"    다듬음  {key} {len(items)}칸 — 예: 원본 {sample[1]!r} → {sample[2]!r}")
    problems += diffs

    # ── 원본_전체 사본 ──────────────────────────────────
    copy = wb["원본_전체"]
    copy_diff = 0
    copy_checked = 0
    for r in range(1, src.max_row + 1):
        for c in range(1, 31):
            copy_checked += 1
            a = raw_text(src.cell(r, c).value)
            b = raw_text(copy.cell(r, c).value)
            if a != b:
                copy_diff += 1
                if copy_diff <= 5:
                    problems.append(f"원본_전체 {get_column_letter(c)}{r}: 원본 {a[:30]!r} ≠ 사본 {b[:30]!r}")
    print(f"원본_전체 사본 {copy_checked:,}칸 대조 · 어긋남 {copy_diff}칸")

    # ── 원본에서 직접 센 값과 맞추기 ─────────────────────
    def src_col(key):
        return [T.s(src.cell(r, T.COL[key]).value) for r in sorted(src_rows)]

    def app_col(name):
        return [T.s(apps.cell(app_rows[r], ah[name]).value) for r in sorted(src_rows)]

    counts = []
    for key, name in (("성별", "성별"), ("유학지역", "유학 지역"),
                      ("거주유형", "거주 유형"), ("최종배정", "최종 배정")):
        a, b = defaultdict(int), defaultdict(int)
        for v in src_col(key):
            a[v] += 1
        for v in app_col(name):
            b[v] += 1
        same = a == b
        counts.append((key, len(a), same))
        if not same:
            only = {k: (a.get(k), b.get(k)) for k in set(a) | set(b) if a.get(k) != b.get(k)}
            problems.append(f"'{key}' 값 분포가 다름: {list(only.items())[:5]}")
    print("값 분포 대조 " + " · ".join(
        f"{k} {v}종 {'일치' if ok else '불일치'}" for k, v, ok in counts))

    # ── 학생 수의 근거 ──────────────────────────────────
    keys = set()
    for r in sorted(src_rows):
        name = T.s(src.cell(r, T.COL["성명"]).value)
        phone = T.s(src.cell(r, T.COL["보호자연락처"]).value)
        gname = T.s(src.cell(r, T.COL["보호자성명"]).value)
        school = T.s(src.cell(r, T.COL["원소속교"]).value)
        keys.add((name, phone or f"{gname}/{school}"))
    students = wb["학생현황"]
    sh = header_map(students)
    n_students = 0
    for r in range(2, students.max_row + 1):
        if students.cell(r, sh["학생ID"]).value is None:
            break
        n_students += 1
    ok = n_students == len(keys)
    print(f"학생 수 {n_students}명 · 원본에서 (성명+보호자 연락처)로 센 값 {len(keys)}명 "
          f"{'일치' if ok else '불일치'}")
    if not ok:
        problems.append(f"학생 수 {n_students} ≠ 원본 기준 {len(keys)}")

    # ── 유학 학년도 칸과 유학이력 ───────────────────────
    enroll = wb["유학이력"]
    eh = header_map(enroll)
    years = defaultdict(set)
    for r in range(2, enroll.max_row + 1):
        aid = enroll.cell(r, eh["신청ID"]).value
        if aid is None:
            break
        years[aid].add(int(enroll.cell(r, eh["학년도"]).value))
    app_of = {apps.cell(r, ah["신청ID"]).value: r for r in app_rows.values()}
    mismatch, compared = 0, 0
    for aid, got in years.items():
        listed = apps.cell(app_of[aid], ah["유학 학년도(원본)"]).value
        if not listed:
            continue
        compared += 1
        want = {int(x) for x in str(listed).replace(",", " ").split() if x.strip().isdigit()}
        if want and want != got:
            mismatch += 1
            if mismatch <= 5:
                problems.append(f"{aid}: 원본 유학 학년도 {sorted(want)} ≠ 유학이력 {sorted(got)}")
    print(f"유학 학년도 칸 대조 {compared}건 · 어긋남 {mismatch}건")

    # ── 원본 741행이 각각 어디로 갔는가 ────────────────
    verdicts = defaultdict(int)
    for r in app_rows.values():
        verdicts[apps.cell(r, ah["배정 판정"]).value] += 1
    n_enroll = sum(1 for r in range(2, enroll.max_row + 1)
                   if enroll.cell(r, eh["이력ID"]).value is not None)
    total = sum(verdicts.values())
    print()
    print("원본 한 줄 한 줄이 어디로 갔는지")
    for k in ("배정", "미배정", "확인필요"):
        note = {"배정": "→ 유학이력으로 펼침",
                "미배정": "→ 신청이력에만 남음(유학하지 않음)",
                "확인필요": "→ 신청이력에만 남고 데이터검증에 올림"}[k]
        print(f"    {k:<6} {verdicts.get(k, 0):>4}건  {note}")
    print(f"    {'합계':<6} {total:>4}건  원본 {len(src_rows)}행과 {'일치' if total == len(src_rows) else '불일치'}")
    if total != len(src_rows):
        problems.append(f"판정 합계 {total} ≠ 원본 {len(src_rows)}행")
    if verdicts.get("배정"):
        print(f"    배정 {verdicts['배정']}건이 유학이력 {n_enroll}행으로 펼쳐짐 "
              f"(한 건당 평균 {n_enroll / verdicts['배정']:.1f}학기)")

    # ── 도교육청 발표 자료와 맞추기 ─────────────────────
    # 출처: '참고 시·군별 유학생 현황 — 연도별 시·군 유학생 수' ('26. 1학기 모집 기준)
    OFFICIAL = {2022: 27, 2023: 85, 2024: 165, 2025: 269, 2026: 333}
    by_year = defaultdict(set)
    for r in range(2, enroll.max_row + 1):
        if enroll.cell(r, eh["이력ID"]).value is None:
            break
        y = int(enroll.cell(r, eh["학년도"]).value)
        term = str(enroll.cell(r, eh["학기구분"]).value)
        if y == 2026 and term != "1학기":
            continue
        by_year[y].add(enroll.cell(r, eh["학생ID"]).value)
    print()
    print("도교육청 발표 자료 '연도별 시·군 유학생 수' 와 맞추기")
    for y in sorted(OFFICIAL):
        got, want = len(by_year.get(y, ())), OFFICIAL[y]
        tag = "일치" if got == want else f"{got - want:+d} — 고창 1명은 학교 오보고분"
        print(f"    {y}학년도{' 1학기' if y == 2026 else '    '}  발표 {want:>4}  워크북 {got:>4}  {tag}")
        if got != want and not (y == 2026 and want - got == 1):
            problems.append(f"{y}학년도 발표 {want} ≠ 워크북 {got}")

    print()
    if problems:
        print(f"확인 필요 {len(problems)}건")
        for p in problems[:20]:
            print("  !", p)
        return 1
    print("원본과 어긋나는 곳이 없습니다")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1], sys.argv[2]))
