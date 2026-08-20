"""전북 농어촌유학생 명단 원데이터 → 정규화 구조로 변환.

원본 워크북의 1번 시트(★☆★전북 농어촌 유학생 명단★☆★)만 원천으로 삼는다.
숨김 시트(연도별 명단, 정보공개 등)는 과거 산출물이므로 읽지 않는다.

산출:
  students     학생 마스터 (1행 = 학생)
  applications 신청 이력   (1행 = 원본 1행 = 모집 차수별 신청 건)
  enrollments  유학 이력   (1행 = 학생 × 학기)
  issues       데이터 검증 항목
"""

from __future__ import annotations

import calendar
import datetime as dt
import re
from collections import defaultdict
from dataclasses import dataclass, field

import openpyxl

SOURCE_SHEET = "★☆★전북 농어촌 유학생 명단★☆★"
HEADER_ROW = 8
FIRST_DATA_ROW = 9

# 기준 학기: '현재' 로 적힌 유학생이 재학 중인 학기.
CURRENT_YEAR, CURRENT_TERM = 2026, 1

# 원본 열 번호(1-based) → 이름
COL = {
    "순": 1,
    "예비유학생": 2,
    "중학교진학": 3,
    "관내전학": 4,
    "기타": 5,
    "유학학년도_원본": 6,
    "접수일": 7,
    "배정희망서": 8,
    "참가신청서": 9,
    "최종배정": 10,
    "중간종료사유": 11,
    "종료일": 12,
    "성명": 13,
    "학생전화": 14,
    "성별": 15,
    "원지역": 16,
    "원소속청": 17,
    "원소속교": 18,
    "유학지역": 19,
    "전입학년": 20,
    "현재학년": 21,
    "유학학교": 22,
    "이전유학지역": 23,
    "이전유학학교": 24,
    "거주유형": 25,
    "거주지": 26,
    "보호자성명": 27,
    "보호자연락처": 28,
    "형제자매": 29,
    "비고": 30,
}

GRADE_ORDER = ["유치원", "초1", "초2", "초3", "초4", "초5", "초6", "중1", "중2", "중3"]
GRADE_INDEX = {g: i for i, g in enumerate(GRADE_ORDER)}

OFFICE_ALIAS = {
    "경기도": "경기",
    "경상남도": "경남",
    "경상북도": "경북",
    "광주광역시": "광주",
    "대구광역시": "대구",
    "대전광역시": "대전",
    "부산광역시": "부산",
    "세종특별시": "세종",
    "울산광역시": "울산",
    "인천광역시": "인천",
    "전라남도": "전남",
    "제주특별자치도": "제주",
    "충청남도": "충남",
}

# 미선정 사유 원문 → 분류. 위에서부터 먼저 맞는 것으로 분류한다.
REASON_RULES = [
    ("거주시설 부족", ["거주시설 부족", "거주시설이 부족", "거주지 부족", "잔여 거주시설"]),
    ("거주시설 부적합·협소", ["협소", "부적합", "인정 불가", "거주공간"]),
    ("거주지 미확보", ["거주지 못", "거주지 미", "거주지를 구하지"]),
    ("거주시설 불만", ["거주시설 불만"]),
    ("면담 불참", ["면담 불참", "면접 불참"]),
    ("면담 결과 미선정", ["면담", "면접"]),
    ("자진 포기·철회", ["자진", "포기", "철회", "미희망"]),
    ("학교 미제출", ["학교 미제출", "미제출"]),
    ("학교생활 적응 우려", ["적응"]),
    ("교육과정 운영 곤란", ["교육과정"]),
    ("중학교 배정 문제", ["중학교"]),
    ("개인 사정", ["개인사", "개인 사정", "공공분양"]),
    ("배정 우선순위 밀림", ["순위", "외동"]),
]

ASSIGNED_TOKENS = {"배정", "최종배정"}
UNASSIGNED_TOKENS = {"X", "x", "미배정"}


def s(v) -> str | None:
    """셀 값을 다듬은 문자열로. 빈 값은 None."""
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    t = str(v).strip()
    t = re.sub(r"[ \t]+", " ", t)
    return t or None


def as_date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    return None


def sem_of(d: dt.date) -> tuple[int, int]:
    """날짜가 속한 (학년도, 학기). 학년도는 3월 시작."""
    if d.month >= 9:
        return d.year, 2
    if d.month >= 3:
        return d.year, 1
    return d.year - 1, 2


def sem_index(year: int, term: int) -> int:
    return year * 2 + (term - 1)


def sem_from_index(i: int) -> tuple[int, int]:
    return i // 2, i % 2 + 1


def sem_label(year: int, term: int) -> str:
    return f"{year}-{term}학기"


def sem_start(year: int, term: int) -> dt.date:
    return dt.date(year, 3, 1) if term == 1 else dt.date(year, 9, 1)


def sem_end(year: int, term: int) -> dt.date:
    if term == 1:
        return dt.date(year, 8, 31)
    last = calendar.monthrange(year + 1, 2)[1]
    return dt.date(year + 1, 2, last)


CURRENT_INDEX = sem_index(CURRENT_YEAR, CURRENT_TERM)


def classify_reason(*texts) -> str | None:
    joined = " ".join(t for t in texts if t)
    if not joined:
        return None
    for label, keys in REASON_RULES:
        if any(k in joined for k in keys):
            return label
    return "기타"


def normalize_grade(v: str | None) -> str | None:
    if not v:
        return None
    v = v.replace(" ", "")
    return v


def grade_after(base: str | None, semesters_passed_years: int) -> str | None:
    """전입 당시 학년에서 학년도가 n번 바뀐 뒤의 학년."""
    if base not in GRADE_INDEX:
        return None
    i = GRADE_INDEX[base] + semesters_passed_years
    if 0 <= i < len(GRADE_ORDER):
        return GRADE_ORDER[i]
    return None


def normalize_residence(raw: str | None, year: int | None = None) -> str | None:
    """거주 유형 표준화. '홈스테이형(2022)→유학센터형(2023)' 같은 값은 학년도로 갈라준다."""
    if not raw:
        return None
    if "→" in raw:
        parts = []
        for chunk in raw.split("→"):
            m = re.match(r"\s*([^\(]+)\((\d{4})\)", chunk)
            if m:
                parts.append((int(m.group(2)), m.group(1).strip()))
        if parts:
            parts.sort()
            if year is None:
                return parts[0][1]
            chosen = parts[0][1]
            for y, kind in parts:
                if year >= y:
                    chosen = kind
            return chosen
        return raw
    return raw


@dataclass
class Application:
    row: int
    raw: dict
    app_id: str = ""
    student_id: str = ""
    intake_date: dt.date | None = None
    intake_year: int | None = None
    intake_term: int | None = None
    intake_round: int | None = None
    term_start: dt.date | None = None
    decision: str = ""
    decision_basis: str = ""
    reject_stage: str | None = None
    reject_reason: str | None = None
    reject_class: str | None = None


@dataclass
class Student:
    student_id: str
    name: str
    gender: str | None = None
    phone: str | None = None
    home_region: str | None = None
    home_office: str | None = None
    home_school: str | None = None
    guardian: str | None = None
    guardian_phone: str | None = None
    household_id: str | None = None
    note: str | None = None
    apps: list[Application] = field(default_factory=list)


@dataclass
class Enrollment:
    student_id: str
    name: str
    year: int
    term: int
    kind: str
    region: str | None
    school: str | None
    grade: str | None
    gender: str | None
    residence: str | None
    place: str | None
    household_id: str | None
    home_region: str | None
    home_school: str | None
    status: str
    end_reason: str | None
    start_date: dt.date | None
    end_date: dt.date | None
    app_id: str
    term_end: dt.date | None = None


def load_rows(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SOURCE_SHEET]
    out = []
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        raw = {name: ws.cell(r, c).value for name, c in COL.items()}
        if not s(raw["성명"]):
            continue
        raw["_row"] = r
        out.append(raw)
    return out


def decide(raw: dict) -> tuple[str, str]:
    """배정 판정과 그 근거."""
    final = s(raw["최종배정"])
    end = s(raw["종료일"])
    wish = s(raw["배정희망서"])

    if end and "미전학" in end:
        return "미배정", f"종료일 칸에 '{end.splitlines()[0]}' — 최종배정 후 전학하지 않음"
    if end in ASSIGNED_TOKENS:
        return "확인필요", (
            f"종료일 칸에 '{end}' 라고만 적혀 있어 실제 전학 여부를 알 수 없음 "
            "— 공식 시군별 집계에도 빠져 있는 건"
        )
    if final in ASSIGNED_TOKENS:
        return "배정", f"최종배정 = {final}"
    if final in UNASSIGNED_TOKENS:
        return "미배정", f"최종배정 = {final}"
    if final:
        return "미배정", f"최종배정 칸에 사유 기재: {final[:40]}"
    if end and (end.startswith("현재") or as_date(raw["종료일"])):
        return "배정", f"최종배정 비어 있으나 종료일({end[:20]}) 기재"
    if wish == "선정":
        return "확인필요", "학교 배정희망서는 '선정'이나 최종배정 결과 미기재"
    if wish:
        return "미배정", f"학교 단계 미선정: {wish[:40]}"
    return "확인필요", "배정 결과 관련 칸이 모두 비어 있음"


def build(path: str):
    rows = load_rows(path)

    # ── 학생 식별: 성명 + 보호자 연락처(없으면 보호자 성명 + 원 소속교)
    students: dict[str, Student] = {}
    order: list[str] = []
    key_to_id: dict[tuple, str] = {}
    apps: list[Application] = []

    for raw in rows:
        name = s(raw["성명"])
        gphone = s(raw["보호자연락처"])
        gname = s(raw["보호자성명"])
        key = (name, gphone or f"{gname}/{s(raw['원소속교'])}")
        if key not in key_to_id:
            sid = f"S{len(key_to_id) + 1:04d}"
            key_to_id[key] = sid
            students[sid] = Student(
                student_id=sid,
                name=name,
                gender=s(raw["성별"]),
                phone=s(raw["학생전화"]),
                home_region=s(raw["원지역"]),
                home_office=OFFICE_ALIAS.get(s(raw["원소속청"]) or "", s(raw["원소속청"])),
                home_school=s(raw["원소속교"]),
                guardian=gname,
                guardian_phone=gphone,
                household_id=None,
                note=s(raw["비고"]),
            )
            order.append(sid)
        sid = key_to_id[key]

        app = Application(row=raw["_row"], raw=raw, student_id=sid)
        app.app_id = f"A{len(apps) + 1:04d}"
        d = as_date(raw["접수일"])
        app.intake_date = d
        if d:
            app.intake_year, app.intake_term = sem_of(d)
            # 접수일 칸의 '일' 자리는 날짜가 아니라 모집 차수다 (2025-09-02 = 2025년 2학기 2차 모집)
            app.intake_round = d.day if d.day <= 9 else 1
            # 차수 자리만 떼고 월은 남긴다 — 2022년 시범사업은 실제로 10월에 시작했다
            app.term_start = d.replace(day=1)
        app.decision, app.decision_basis = decide(raw)
        if app.decision != "배정":
            wish, submit, final = (
                s(raw["배정희망서"]),
                s(raw["참가신청서"]),
                s(raw["최종배정"]),
            )
            if final and final not in UNASSIGNED_TOKENS:
                app.reject_stage, app.reject_reason = "도교육청", final
            elif submit and submit != "제출":
                app.reject_stage, app.reject_reason = "학부모", submit
            elif wish and wish != "선정":
                app.reject_stage, app.reject_reason = "학교", wish
            elif final in UNASSIGNED_TOKENS:
                app.reject_stage, app.reject_reason = "도교육청", "미배정(사유 미기재)"
            app.reject_class = classify_reason(app.reject_reason)
        apps.append(app)
        students[sid].apps.append(app)

    # ── 가구: 보호자 연락처 기준
    house_of: dict[str, str] = {}
    for sid in order:
        st = students[sid]
        hkey = st.guardian_phone or f"{st.guardian}/{st.home_school}"
        if hkey not in house_of:
            house_of[hkey] = f"H{len(house_of) + 1:04d}"
        st.household_id = house_of[hkey]

    # ── 유학 이력: 배정 건을 학기 단위로 펼친다
    enrollments: list[Enrollment] = []
    issues: list[dict] = []

    def add_issue(kind, level, app: Application | None, detail):
        issues.append(
            {
                "유형": kind,
                "심각도": level,
                "원본행": app.row if app else None,
                "신청ID": app.app_id if app else None,
                "학생ID": app.student_id if app else None,
                "성명": s(app.raw["성명"]) if app else None,
                "내용": detail,
            }
        )

    for app in apps:
        raw = app.raw
        if app.intake_date is None:
            add_issue("접수일 없음", "높음", app, "접수일(시작일)이 비어 있어 학기를 계산할 수 없음")
            continue
        if app.decision == "확인필요":
            kind = ("최종배정 후 전학 확인 필요"
                    if s(raw["종료일"]) in ASSIGNED_TOKENS else "배정 결과 미기재")
            add_issue(kind, "높음", app, app.decision_basis)
            continue
        if app.decision != "배정":
            continue

        start_i = sem_index(app.intake_year, app.intake_term)
        end_raw = s(raw["종료일"])
        end_date = as_date(raw["종료일"])
        ongoing = False
        if end_date:
            end_i = sem_index(*sem_of(end_date))
        elif end_raw and end_raw.startswith("현재"):
            end_i, ongoing = max(start_i, CURRENT_INDEX), True
        else:
            end_i, ongoing = max(start_i, CURRENT_INDEX), True
            add_issue(
                "종료일 미기재",
                "보통",
                app,
                f"배정 건이나 종료일 칸이 '{end_raw or '(비어 있음)'}' — 재학 중으로 간주함",
            )
        if end_i < start_i:
            add_issue(
                "종료일 역전",
                "높음",
                app,
                f"종료일({end_raw})이 시작 학기({sem_label(app.intake_year, app.intake_term)})보다 이름 — 1개 학기로 처리",
            )
            end_i = start_i

        base_grade = normalize_grade(s(raw["전입학년"]))
        if base_grade and base_grade not in GRADE_INDEX:
            add_issue("학년 표기 오류", "보통", app, f"전입시 학년 '{base_grade}' 은 표준 학년 값이 아님")
        region = (s(raw["유학지역"]) or "").strip() or None
        school = s(raw["유학학교"])
        end_reason = s(raw["중간종료사유"])

        for i in range(start_i, end_i + 1):
            y, t = sem_from_index(i)
            if i > CURRENT_INDEX:
                status = "예정"
            elif i == end_i and not ongoing:
                status = "종료"
            else:
                status = "재학"
            enrollments.append(
                Enrollment(
                    student_id=app.student_id,
                    name=s(raw["성명"]),
                    year=y,
                    term=t,
                    kind="",
                    region=region,
                    school=school,
                    grade=grade_after(base_grade, y - app.intake_year),
                    gender=s(raw["성별"]),
                    residence=normalize_residence(s(raw["거주유형"]), y),
                    place=s(raw["거주지"]),
                    household_id=students[app.student_id].household_id,
                    home_region=s(raw["원지역"]),
                    home_school=s(raw["원소속교"]),
                    status=status,
                    end_reason=end_reason if (i == end_i and not ongoing) else None,
                    start_date=sem_start(y, t) if i > start_i else app.term_start,
                    end_date=sem_end(y, t) if i < end_i else (end_date or None),
                    app_id=app.app_id,
                    term_end=sem_end(y, t),
                )
            )

        # 원본에 적힌 '유학 학년도' 와 계산 결과가 맞는지 대조
        listed = s(raw["유학학년도_원본"])
        if listed:
            want = {int(x) for x in re.findall(r"\d{4}", listed)}
            got = {sem_from_index(i)[0] for i in range(start_i, end_i + 1)}
            if want != got:
                add_issue(
                    "유학 학년도 불일치",
                    "보통",
                    app,
                    f"원본 기재 {sorted(want)} ↔ 접수일·종료일로 계산 {sorted(got)}",
                )

    # 학생별 학기 정렬 → 신규/계속/재유학
    by_student: dict[str, list[Enrollment]] = defaultdict(list)
    for e in enrollments:
        by_student[e.student_id].append(e)
    for sid, lst in by_student.items():
        lst.sort(key=lambda e: sem_index(e.year, e.term))
        prev = None
        for e in lst:
            idx = sem_index(e.year, e.term)
            if prev is None:
                e.kind = "신규"
            elif idx == prev + 1:
                e.kind = "계속"
            else:
                e.kind = "재유학"
            prev = idx

    enrollments.sort(key=lambda e: (sem_index(e.year, e.term), e.region or "", e.school or "", e.name))

    # 현재 학년 대조: 재학 중인 학생의 기준학기 학년이 원본 '현재 학년' 과 같은지
    cur_grade = {
        e.student_id: e.grade
        for e in enrollments
        if sem_index(e.year, e.term) == CURRENT_INDEX
    }
    for app in apps:
        if app.decision != "배정":
            continue
        listed = normalize_grade(s(app.raw["현재학년"]))
        got = cur_grade.get(app.student_id)
        if listed and got and listed != got:
            issues.append(
                {
                    "유형": "현재 학년 불일치",
                    "심각도": "낮음",
                    "원본행": app.row,
                    "신청ID": app.app_id,
                    "학생ID": app.student_id,
                    "성명": s(app.raw["성명"]),
                    "내용": f"원본 '현재 학년' {listed} ↔ 전입 학년 기준 계산 {got}",
                }
            )

    # 한 학생이 같은 학기에 두 번 재학 중으로 잡히는 경우(전학 처리 후 이전 행을 닫지 않은 등)
    seen: dict[tuple, Enrollment] = {}
    for e in enrollments:
        k = (e.student_id, e.year, e.term)
        if k in seen:
            other = seen[k]
            issues.append(
                {
                    "유형": "같은 학기 중복 재학",
                    "심각도": "높음",
                    "원본행": None,
                    "신청ID": f"{other.app_id}, {e.app_id}",
                    "학생ID": e.student_id,
                    "성명": e.name,
                    "내용": (
                        f"{sem_label(e.year, e.term)}에 {other.region} {other.school} 과 "
                        f"{e.region} {e.school} 두 곳으로 잡힘 — 이전 신청 건의 종료일 기재 필요"
                    ),
                }
            )
        else:
            seen[k] = e

    # 한 사람으로 묶었으나 원 소속교가 다른 경우 — 이사 후 재신청일 수도, 동명이인일 수도 있다
    for sid in order:
        st = students[sid]
        schools = {s(a.raw["원소속교"]) for a in st.apps if s(a.raw["원소속교"])}
        if len(schools) > 1:
            issues.append(
                {
                    "유형": "동일인 여부 확인",
                    "심각도": "보통",
                    "원본행": ", ".join(str(a.row) for a in st.apps),
                    "신청ID": ", ".join(a.app_id for a in st.apps),
                    "학생ID": sid,
                    "성명": st.name,
                    "내용": (
                        "이름과 보호자 연락처가 같아 한 사람으로 묶었으나 원 소속교가 다름 "
                        f"({' / '.join(sorted(schools))}) — 이사 후 재신청인지 동명이인인지 확인 필요"
                    ),
                }
            )

    # 동명이인 경고
    name_map: dict[str, set] = defaultdict(set)
    for sid in order:
        name_map[students[sid].name].add(sid)
    for name, sids in sorted(name_map.items()):
        if len(sids) > 1:
            issues.append(
                {
                    "유형": "동명이인",
                    "심각도": "낮음",
                    "원본행": None,
                    "신청ID": None,
                    "학생ID": ", ".join(sorted(sids)),
                    "성명": name,
                    "내용": f"같은 이름이 {len(sids)}명 — 보호자 연락처로 구분됨",
                }
            )

    issues.sort(key=lambda x: ({"높음": 0, "보통": 1, "낮음": 2}[x["심각도"]], x["유형"], x["원본행"] or 0))

    semesters = sorted({sem_index(e.year, e.term) for e in enrollments})
    return {
        "students": [students[sid] for sid in order],
        "applications": apps,
        "enrollments": enrollments,
        "issues": issues,
        "by_student": by_student,
        "semesters": [sem_from_index(i) for i in semesters],
    }
