"""워크북에 쓴 수식을 실제로 계산하는 작은 계산기.

LibreOffice 를 못 쓰는 환경에서 재계산을 대신한다. 이 워크북이 쓰는 함수
(COUNTIF, COUNTIFS, SUM, IF, OR, MIN, MAX, INDEX)만 다룬다.
"""

from __future__ import annotations

import datetime as dt
import re

from openpyxl.utils import column_index_from_string, get_column_letter

TOKEN_RE = re.compile(
    r"""
    (?P<ws>\s+)
  | (?P<string>"(?:[^"]|"")*")
  | (?P<sheet>'[^']+'!|[가-힣A-Za-z_][가-힣\w]*!)
  | (?P<ref>\$?[A-Z]{1,3}\$?\d+)
  | (?P<number>\d+(?:\.\d+)?)
  | (?P<op><>|>=|<=|[=<>+\-*/&:,()])
  | (?P<name>[A-Z][A-Z0-9_.]*)
    """,
    re.VERBOSE,
)


class Err(Exception):
    pass


def tokenize(text: str):
    pos, out = 0, []
    while pos < len(text):
        m = TOKEN_RE.match(text, pos)
        if not m:
            raise Err(f"읽을 수 없는 글자: {text[pos:pos + 12]!r}")
        pos = m.end()
        kind = m.lastgroup
        if kind != "ws":
            out.append((kind, m.group()))
    return out


class Range:
    def __init__(self, sheet, col1, row1, col2, row2):
        self.sheet, self.col1, self.row1, self.col2, self.row2 = sheet, col1, row1, col2, row2

    def cells(self):
        c1, c2 = column_index_from_string(self.col1), column_index_from_string(self.col2)
        for c in range(min(c1, c2), max(c1, c2) + 1):
            for r in range(min(self.row1, self.row2), max(self.row1, self.row2) + 1):
                yield self.sheet, f"{get_column_letter(c)}{r}"


class Parser:
    def __init__(self, tokens, book, sheet, coord):
        self.t, self.i = tokens, 0
        self.book, self.sheet, self.coord = book, sheet, coord

    def peek(self):
        return self.t[self.i] if self.i < len(self.t) else (None, None)

    def take(self, value=None):
        kind, text = self.peek()
        if value is not None and text != value:
            raise Err(f"{value!r} 를 기대했는데 {text!r}")
        self.i += 1
        return text

    def parse(self):
        v = self.comparison()
        if self.i != len(self.t):
            raise Err(f"남은 토큰: {self.t[self.i:]}")
        return v

    def comparison(self):
        left = self.arith()
        kind, text = self.peek()
        if text in ("=", "<>", ">", "<", ">=", "<="):
            self.take()
            right = self.arith()
            return compare(left, text, right)
        return left

    def arith(self):
        v = self.term()
        while self.peek()[1] in ("+", "-", "&"):
            op = self.take()
            r = self.term()
            if op == "&":
                v = f"{as_text(v)}{as_text(r)}"
            else:
                v = num(v) + num(r) if op == "+" else num(v) - num(r)
        return v

    def term(self):
        v = self.unary()
        while self.peek()[1] in ("*", "/"):
            op = self.take()
            r = self.unary()
            if op == "*":
                v = num(v) * num(r)
            else:
                d = num(r)
                if d == 0:
                    raise Err("0 으로 나눔")
                v = num(v) / d
        return v

    def unary(self):
        if self.peek()[1] == "-":
            self.take()
            return -num(self.atom())
        return self.atom()

    def atom(self):
        kind, text = self.peek()
        if kind == "number":
            self.take()
            return float(text)
        if kind == "string":
            self.take()
            return text[1:-1].replace('""', '"')
        if text == "(":
            self.take()
            v = self.comparison()
            self.take(")")
            return v
        if kind == "name":
            return self.call()
        if kind in ("sheet", "ref"):
            return self.reference()
        raise Err(f"뜻밖의 토큰 {text!r}")

    def reference(self):
        kind, text = self.peek()
        sheet = self.sheet
        if kind == "sheet":
            self.take()
            sheet = text[:-1].strip("'")
        col1, row1 = split_ref(self.take())
        if self.peek()[1] == ":":
            self.take()
            if self.peek()[0] == "sheet":
                self.take()
            col2, row2 = split_ref(self.take())
            return Range(sheet, col1, row1, col2, row2)
        return self.book.value(sheet, f"{col1}{row1}")

    def skip_arg(self):
        """인자 하나를 계산하지 않고 건너뛴다."""
        depth = 0
        while self.i < len(self.t):
            text = self.t[self.i][1]
            if depth == 0 and text in (",", ")"):
                return
            if text == "(":
                depth += 1
            elif text == ")":
                depth -= 1
            self.i += 1

    def call_if(self):
        """IF 는 필요한 가지만 계산한다. 엑셀도 그렇게 하고, 그래야 0 으로 나누기를 피한다."""
        cond = truthy(self.comparison())
        result = False
        if self.peek()[1] == ",":
            self.take()
            if cond:
                result = self.comparison()
            else:
                self.skip_arg()
        if self.peek()[1] == ",":
            self.take()
            if cond:
                self.skip_arg()
            else:
                result = self.comparison()
        self.take(")")
        return result

    def call(self):
        name = self.take().upper()
        self.take("(")
        if name == "IF":
            return self.call_if()
        args = []
        if self.peek()[1] != ")":
            while True:
                args.append(self.comparison())
                if self.peek()[1] == ",":
                    self.take()
                    continue
                break
        self.take(")")
        return apply_function(name, args, self.book)


def split_ref(text):
    m = re.fullmatch(r"\$?([A-Z]{1,3})\$?(\d+)", text)
    if not m:
        raise Err(f"셀 주소가 아님: {text!r}")
    return m.group(1), int(m.group(2))


def num(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, bool):
        return 1.0 if v else 0.0
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, (dt.date, dt.datetime)):
        return float(v.toordinal())
    try:
        return float(v)
    except (TypeError, ValueError):
        raise Err(f"숫자가 아님: {v!r}")


def as_text(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def compare(left, op, right):
    if isinstance(left, (int, float)) and not isinstance(left, bool) or isinstance(
        right, (int, float)
    ) and not isinstance(right, bool):
        try:
            a, b = num(left), num(right)
        except Err:
            a, b = as_text(left).lower(), as_text(right).lower()
    elif isinstance(left, (dt.date, dt.datetime)) or isinstance(right, (dt.date, dt.datetime)):
        a, b = to_date(left), to_date(right)
        if a is None or b is None:
            a, b = as_text(left), as_text(right)
    else:
        a, b = as_text(left).lower(), as_text(right).lower()
    return {
        "=": a == b, "<>": a != b, ">": a > b, "<": a < b, ">=": a >= b, "<=": a <= b
    }[op]


def to_date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    return None


CRITERION_RE = re.compile(r"^(<>|>=|<=|>|<|=)?(.*)$", re.DOTALL)


def match_criterion(value, criterion):
    if isinstance(criterion, (int, float)) and not isinstance(criterion, bool):
        try:
            return num(value) == float(criterion)
        except Err:
            return False
    text = as_text(criterion)
    op, rest = CRITERION_RE.match(text).groups()
    if op is None or op == "=":
        if rest == "":
            return value is None or value == ""
        return as_text(value).lower() == rest.lower()
    if op == "<>":
        if rest == "":
            return not (value is None or value == "")
        return as_text(value).lower() != rest.lower()
    if value is None or value == "":
        return False
    try:
        return compare(num(value), op, float(rest))
    except Err:
        return compare(value, op, rest)


def apply_function(name, args, book):
    def values(a):
        return [book.value(s, c) for s, c in a.cells()] if isinstance(a, Range) else [a]

    if name == "SUM":
        total = 0.0
        for a in args:
            for v in values(a):
                if v is None or v == "":
                    continue
                try:
                    total += num(v)
                except Err:
                    pass
        return total
    if name == "COUNTA":
        return float(sum(1 for a in args for v in values(a) if v not in (None, "")))
    if name == "COUNTBLANK":
        return float(sum(1 for a in args for v in values(a) if v in (None, "")))
    if name == "COUNTIF":
        return float(sum(1 for v in values(args[0]) if match_criterion(v, args[1])))
    if name == "COUNTIFS":
        pairs = [(values(args[i]), args[i + 1]) for i in range(0, len(args), 2)]
        n = len(pairs[0][0])
        return float(
            sum(
                1
                for i in range(n)
                if all(match_criterion(vals[i], crit) for vals, crit in pairs)
            )
        )
    if name == "IF":
        return args[1] if truthy(args[0]) else (args[2] if len(args) > 2 else False)
    if name == "OR":
        return any(truthy(a) for a in args)
    if name == "AND":
        return all(truthy(a) for a in args)
    if name == "MIN":
        return min(num(v) for a in args for v in values(a))
    if name == "MAX":
        return max(num(v) for a in args for v in values(a))
    if name == "INDEX":
        if len(args) != 2:
            raise Err(f"INDEX 인자가 {len(args)}개 — 범위와 위치 두 개여야 함")
        cells = list(args[0].cells())
        i = int(num(args[1]))
        if not 1 <= i <= len(cells):
            raise Err(f"INDEX 위치 {i} 가 범위({len(cells)}칸) 밖")
        s, c = cells[i - 1]
        return book.value(s, c)
    if name == "IFERROR":
        return args[0]
    raise Err(f"모르는 함수 {name}")


def truthy(v):
    if isinstance(v, bool):
        return v
    if v is None or v == "":
        return False
    try:
        return num(v) != 0
    except Err:
        return True


class Book:
    """수식 셀을 필요할 때 계산해 주는 워크북."""

    def __init__(self, wb):
        self.cells = {}
        for ws in wb.worksheets:
            sheet = {}
            for row in ws.iter_rows():
                for c in row:
                    if c.value is not None:
                        sheet[c.coordinate] = c.value
            self.cells[ws.title] = sheet
        self.cache = {}
        self.stack = set()
        self.errors = []

    def value(self, sheet, coord):
        key = (sheet, coord)
        if key in self.cache:
            return self.cache[key]
        raw = self.cells.get(sheet, {}).get(coord)
        if not (isinstance(raw, str) and raw.startswith("=")):
            return raw
        if key in self.stack:
            raise Err(f"순환 참조 {sheet}!{coord}")
        self.stack.add(key)
        try:
            v = Parser(tokenize(raw[1:]), self, sheet, coord).parse()
        except Err as e:
            self.errors.append(f"{sheet}!{coord}: {e}  <- {raw[:90]}")
            v = None
        finally:
            self.stack.discard(key)
        self.cache[key] = v
        return v
